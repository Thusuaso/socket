from api.sql import SqlConnect
from model.financedetail import *
from openpyxl import *
import shutil
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class FinanceTest:
    def __init__(self):
        self.sql = SqlConnect().data
        self.customers = self.sql.getList("""
                                                select 
                                                    m.ID,
                                                    m.FirmaAdi,
                                                    m.Marketing
                                                from MusterilerTB m 
                                                where m.ID in 
                                                    (
												        select 
                                                            s.MusteriID 
                                                        from SiparislerTB s 
                                                        where 
                                                            s.MusteriID = m.ID
                                                    )
                                                    and m.Mt_No=2 and m.ID not in (6,34,43,314,153) and m.Marketing in ('Mekmar','Imperial Homes')
                                          """)
        self.customersFilter = self.sql.getList("""
select 
                                                    m.ID,
                                                    m.FirmaAdi,
                                                    m.Marketing
                                                from MusterilerTB m 
                                                where m.ID in 
                                                    (
												        select 
                                                            s.MusteriID 
                                                        from SiparislerTB s 
                                                        where 
                                                            s.MusteriID = m.ID
                                                    )
                                                    and m.Mt_No=2 and m.ID not in (6,34,43,314,153) and m.Marketing in ('İç Piyasa','Mekmer','Mekmar')
                                          """)
        
        
        
        self.orders = self.sql.getList("""
                                            select 
                                                sum(s.NavlunSatis) as Navlun,
                                                sum(s.DetayTutar_1) as Detay1,
                                                sum(s.DetayTutar_2) as Detay2,
                                                sum(s.DetayTutar_3) as Detay3,
                                                sum(s.sigorta_tutar_satis) as Sigorta,
                                                s.MusteriID,
                                                (select k.KullaniciAdi from KullaniciTB k where k.ID = m.MusteriTemsilciId) as Temsilci
                                            from SiparislerTB s
                                                inner join MusterilerTB m on m.ID = s.MusteriID 
                                            where
                                                m.Mt_No = 2 and s.SiparisDurumID in (1,2,3)
                                            group by
                                                s.MusteriID,m.MusteriTemsilciId
                                       """)
        self.products = self.sql.getList("""
                                            select 
                                                sum(su.SatisToplam) as SatisToplam,
                                                s.MusteriID
                                            from SiparislerTB s 
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
											where
												s.MusteriID in (
																	select m.ID from MusterilerTB m where m.ID = s.MusteriID and m.Mt_No=2
															   )
                                            group by
                                                s.MusteriID
                                         """)
        
        self.productsFilter = self.sql.getList("""
                                                                                            select 
                                                SUM(su.AlisFiyati * su.Miktar) as SatisToplam,
                                                s.MusteriID
                                            from SiparislerTB s 
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
											where
												s.MusteriID in (
																	select m.ID from MusterilerTB m where m.ID = s.MusteriID and m.Mt_No=2
															   ) and su.TedarikciID in (1,123)
                                            group by
                                                s.MusteriID
                                               """)
        
        self.paids = self.sql.getList("""
                                            select 
                                                sum(o.Tutar) as Tutar,
                                                s.MusteriID

                                            from OdemelerTB o 

                                            inner join SiparislerTB s on s.SiparisNo = o.SiparisNo

                                            where
                                                s.MusteriID in (
                                                                select m.ID from MusterilerTB m where m.Mt_No=2
                                                        )
                                            group by
                                                s.MusteriID
                                      """)
        
        self.paidsFilter = self.sql.getList("""
                                            select 
                                                sum(o.Tutar) as Tutar,
                                                s.MusteriID

                                            from Odemeler_MekmerTB o 

                                            inner join SiparislerTB s on s.SiparisNo = o.SiparisNo

                                            where
                                                s.MusteriID in (
                                                                select m.ID from MusterilerTB m where m.Mt_No=2
                                                        )
                                            group by
                                                s.MusteriID
                                      """)
        
        
        
        
        
        
        
        self.paidsForwarding = self.sql.getList("""
                                            select 
                                                sum(o.Tutar) as Tutar,
                                                s.MusteriID

                                            from OdemelerTB o 

                                            inner join SiparislerTB s on s.SiparisNo = o.SiparisNo

                                            where
                                                s.MusteriID in (
                                                                select m.ID from MusterilerTB m where m.Mt_No=2
                                                        )
                                                and s.SiparisDurumID = 3
                                            group by
                                                s.MusteriID
                                      """)
        self.paidsProduction = self.sql.getList("""
                                            select 
                                                sum(o.Tutar) as Tutar,
                                                s.MusteriID

                                            from OdemelerTB o 

                                            inner join SiparislerTB s on s.SiparisNo = o.SiparisNo

                                            where
                                                s.MusteriID in (
                                                                select m.ID from MusterilerTB m where m.Mt_No=2
                                                        )
                                                and s.SiparisDurumID in (1,2)
                                            group by
                                                s.MusteriID
                                      """)
        
        self.ordersForwarding = self.sql.getList("""
                                                        select 
                                                            sum(s.NavlunSatis) as Navlun,
                                                            sum(s.DetayTutar_1) as Detay1,
                                                            sum(s.DetayTutar_2) as Detay2,
                                                            sum(s.DetayTutar_3) as Detay3,
                                                            sum(s.sigorta_tutar_satis) as Sigorta,
                                                            s.MusteriID,
                                                            (select k.KullaniciAdi from KullaniciTB k where k.ID = m.MusteriTemsilciId) as Temsilci
                                                        from SiparislerTB s
                                                            inner join MusterilerTB m on m.ID = s.MusteriID 
                                                        where
                                                            m.Mt_No = 2 and s.SiparisDurumID = 3
                                                        group by
                                                            s.MusteriID,m.MusteriTemsilciId
                                                 """)
        
        
        
        
        
        self.ordersProduction = self.sql.getList("""
                                                    select 
                                                        sum(s.NavlunSatis) as Navlun,
                                                        sum(s.DetayTutar_1) as Detay1,
                                                        sum(s.DetayTutar_2) as Detay2,
                                                        sum(s.DetayTutar_3) as Detay3,
                                                        sum(s.sigorta_tutar_satis) as Sigorta,
                                                        s.MusteriID,
                                                        (select k.KullaniciAdi from KullaniciTB k where k.ID = m.MusteriTemsilciId) as Temsilci
                                                    from SiparislerTB s
                                                        inner join MusterilerTB m on m.ID = s.MusteriID 
                                                    where
                                                        m.Mt_No = 2 and s.SiparisDurumID in (1,2)
                                                    group by
                                                        s.MusteriID,m.MusteriTemsilciId
                                                 
                                                 """)
        self.productsForwarding = self.sql.getList("""
                                                        select 
                                                            sum(su.SatisToplam) as SatisToplam,
                                                            s.MusteriID
                                                        from SiparislerTB s 
                                                            inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                        where
                                                            s.MusteriID in (
                                                                                select m.ID from MusterilerTB m where m.ID = s.MusteriID and m.Mt_No=2
                                                                        )
                                                            and s.SiparisDurumID=3
                                                        group by
                                                            s.MusteriID
                                                   
                                                   """)
        
        self.productsForwardingFilter = self.sql.getList("""
                                                        select 
                                                            sum(su.AlisFiyati * su.Miktar) as SatisToplam,
                                                            s.MusteriID
                                                        from SiparislerTB s 
                                                            inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                        where
                                                            s.MusteriID in (
                                                                                select m.ID from MusterilerTB m where m.ID = s.MusteriID and m.Mt_No=2
                                                                        )
                                                            and s.SiparisDurumID=3 and su.TedarikciID in (1,123)
                                                        group by
                                                            s.MusteriID
                                                   
                                                   """)

        
        
        self.productsProduction = self.sql.getList("""
                                                        select 
                                                            sum(su.SatisToplam) as SatisToplam,
                                                            s.MusteriID
                                                        from SiparislerTB s 
                                                            inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                        where
                                                            s.MusteriID in (
                                                                                select m.ID from MusterilerTB m where m.ID = s.MusteriID and m.Mt_No=2
                                                                        )
                                                            and s.SiparisDurumID in (1,2)
                                                        group by
                                                            s.MusteriID
                                                   
                                                   """)
        self.productsProductionFilter = self.sql.getList("""
                                                        select 
                                                            sum(su.AlisFiyati * su.Miktar) as SatisToplam,
                                                            s.MusteriID
                                                        from SiparislerTB s 
                                                            inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                        where
                                                            s.MusteriID in (
                                                                                select m.ID from MusterilerTB m where m.ID = s.MusteriID and m.Mt_No=2
                                                                        )
                                                            and s.SiparisDurumID in (1,2) and su.TedarikciID in (1,123)
                                                        group by
                                                            s.MusteriID
                                                   
                                                   """)
        
        self.advancePayment = self.sql.getList("""
                                                    select 
                                                        sum(s.Pesinat) as Pesinat,
                                                        s.MusteriID,
                                                        (select k.KullaniciAdi from KullaniciTB k where k.ID = m.MusteriTemsilciId) as Temsilci
                                                    from SiparislerTB s
                                                        inner join MusterilerTB m on m.ID = s.MusteriID 
                                                    where
                                                        m.Mt_No = 2 and s.SiparisDurumID in (1,2)
                                                    group by
                                                        s.MusteriID,m.MusteriTemsilciId
                                               
                                               """)
        
        self.advancePaymentFilter = self.sql.getList("""
                                                    select 
                                                        sum(s.Pesinat) as Pesinat,
                                                        s.MusteriID,
                                                        (select k.KullaniciAdi from KullaniciTB k where k.ID = m.MusteriTemsilciId) as Temsilci
                                                    from SiparislerTB s
                                                        inner join MusterilerTB m on m.ID = s.MusteriID 
                                                    where
                                                        m.Mt_No = 2 and s.SiparisDurumID in (1,2) and m.Marketing in ('İç Piyasa','Imperial Homes')
                                                    group by
                                                        s.MusteriID,m.MusteriTemsilciId
                                               
                                               """)
        
        self.mayaPaymentOrders = self.sql.getList("""
                                                    select 
                                                        s.SiparisNo,
                                                        s.NavlunSatis + s.DetayTutar_1 + s.DetayTutar_2 + s.DetayTutar_3 + s.sigorta_tutar_satis as Cost,
                                                        m.FirmaAdi,
                                                        m.ID,
                                                        s.SiparisTarihi,
														s.YuklemeTarihi
                                                    from SiparislerTB s
                                                        inner join MusterilerTB m on m.ID = s.MusteriID

                                                    where
                                                        s.MayaControl = 1
                                                 """)
        
        self.mayaPaymentProducts = self.sql.getList("""
                                                        select sum(su.SatisToplam) as Total,su.SiparisNo from SiparisUrunTB su
                                                            inner join SiparislerTB s on s.SiparisNo = su.SiparisNo
                                                        where
                                                            s.MayaControl = 1
                                                        group by
                                                            su.SiparisNo
                                                    """)
        self.mayaPaymentList = self.sql.getList("""
                                                    select 
                                                        sum(o.Tutar) as Total,
                                                        o.SiparisNo
                                                    from OdemelerTB o 
                                                        inner join SiparislerTB s on s.SiparisNo = o.SiparisNo
                                                    where
                                                        s.MayaControl = 1

                                                    group by
                                                        o.SiparisNo

                                                """)
        
        
        
        
    def getList(self):
        liste = list()
        for item in self.customers:
            # total = (self.__getOrder(item.ID) + self.__getProduct(item.ID)) - self.__noneControl(self.__getPaid(item.ID))
            # if total == 0 or total < 8:
            #     continue
            # else:
                
            liste.append({
                'marketing':item.Marketing,
                'customer_id':item.ID,
                'customer_name':item.FirmaAdi,
                'order':self.__getOrder(item.ID),
                'product':self.__getProduct(item.ID),
                'total_order_amount': self.__noneControl(self.__getOrder(item.ID)) + self.__noneControl(self.__getProduct(item.ID)),
                'paid':self.__getPaid(item.ID),
                'forwarding': self.__noneControl(self.__getOrderForwarding(item.ID)) + self.__noneControl(self.__getProductForwarding(item.ID)),
                'production':self.__noneControl(self.__getProductProduction(item.ID)) + self.__noneControl(self.__getOrderProduction(item.ID)),
                'advanced_payment':self.__getAdvancePayment(item.ID),
                'total': (self.__getOrder(item.ID) + self.__getProduct(item.ID)) - self.__noneControl(self.__getPaid(item.ID))
            })
        return liste
    
    
    def getListFilter(self):
        liste = list()
        for item in self.customersFilter:

                
            total = (self.__noneControl(self.__getProductForwardingFilter(item.ID)) + self.__noneControl(self.__getProductProductionFilter(item.ID))) - self.__noneControl(self.__getPaidFilter(item.ID))

                
            if total == 0 or total < 8:

                continue
            else:
                liste.append({
                    'marketing':item.Marketing,
                    'customer_id':item.ID,
                    'customer_name':item.FirmaAdi,
                    'order':self.__getOrder(item.ID),
                    'product':self.__getProductFilter(item.ID),
                    'total_order_amount':  self.__noneControl(self.__getProductFilter(item.ID)),
                    'paid':self.__getPaidFilter(item.ID),
                    
                    
                    
                    'forwarding':  self.__noneControl(self.__getProductForwardingFilter(item.ID)),
                    'production':self.__noneControl(self.__getProductProductionFilter(item.ID)),
                    'advanced_payment':self.__getAdvancePaymentFilter(item.ID),
                    'total':  (self.__noneControl(self.__getProductForwardingFilter(item.ID)) + self.__noneControl(self.__getProductProductionFilter(item.ID))) - self.__noneControl(self.__getPaidFilter(item.ID))
                })
        
        return liste
    
    
    def __getProductFilter(self,customer_id):
        for item in self.productsFilter:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.SatisToplam)
    
    def getPoPaidList(self,po):
        try:
            data = self.sql.getStoreList("""select omt.*,(select m.FirmaAdi from MusterilerTB m where m.ID = omt.MusteriID) as MusteriAdi from Odemeler_MekmerTB omt where SiparisNo=?""",po)
            liste = list()
            for item in  data:
                model = MusteriOdemeSecimModel()
                model.id = item.ID
                model.siparisno = item.SiparisNo
                model.tutar = item.Tutar
                model.aciklama = item.Aciklama
                model.masraf = item.Masraf
                model.kur = item.Kur
                model.tarih = item.Tarih
                model.musteri_id = item.MusteriID
                model.musteriadi = item.MusteriAdi
  
                liste.append(model)
            schema = MusteriOdemeSecimSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getPoPaidList hata',str(e))
            return False
    
    def getMayaList(self):
        try:
            liste = list()
            for item in self.mayaPaymentOrders:
                liste.append({
                    'po':item.SiparisNo,
                    'customer':item.FirmaAdi,
                    'order_date':item.SiparisTarihi,
                    'forwarding_date':item.YuklemeTarihi,
                    'order_amount': self.__noneControl(item.Cost) + self.__noneControl(self.__getMayaProduct(item.SiparisNo)),
                    'paid':self.__noneControl(self.__getMayaPaid(item.SiparisNo)),
                    'balance': (self.__noneControl(item.Cost) + self.__noneControl(self.__getMayaProduct(item.SiparisNo))) - self.__noneControl(self.__getMayaPaid(item.SiparisNo))
                })
            return liste
        except Exception as e:
            print('finance maya hata',str(e))
            return False
    
    def getExcelList(self,data_list):
        try:
            source_path = 'excel/sablonlar/finans_test_list.xlsx'
            target_path = 'excel/dosyalar/finans_test_list.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            satir = 2
            new_list = sorted(data_list, key=lambda x: x['balanced'], reverse=True)
            for item in new_list:

                sayfa.cell(satir,column=1,value=item['customer_name'])               
                sayfa.cell(satir,column=2,value=self.__noneControl(item['total_order_amount']))
                sayfa.cell(satir,column=3,value=self.__noneControl(item['production']))
                sayfa.cell(satir,column=4,value=self.__noneControl(item['forwarding']))
                sayfa.cell(satir,column=5,value=self.__noneControl(item['paid']))
                sayfa.cell(satir,column=6,value=self.__noneControl(item['advanced_payment']))
                sayfa.cell(satir,column=7,value=self.__noneControl(item['total']))
                sayfa.cell(satir,column=8,value=(self.__noneControl(item['balanced'])))
                
                

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False
    
    
    def getExcelListMekmer(self,data_list):
        try:
            source_path = 'excel/sablonlar/finans_test_list.xlsx'
            target_path = 'excel/dosyalar/finans_test_list.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            satir = 2
            new_list = sorted(data_list, key=lambda x: x['balanced'], reverse=True)
            for item in new_list:

                sayfa.cell(satir,column=1,value=item['customer_name'])               
                sayfa.cell(satir,column=2,value=self.__noneControl(item['total_order_amount']))
                sayfa.cell(satir,column=3,value=self.__noneControl(item['production']))
                sayfa.cell(satir,column=4,value=self.__noneControl(item['forwarding']))
                sayfa.cell(satir,column=5,value=self.__noneControl(item['paid']))
                sayfa.cell(satir,column=6,value=self.__noneControl(item['advanced_payment']))
                sayfa.cell(satir,column=7,value=self.__noneControl(item['total']))
                sayfa.cell(satir,column=8,value=(self.__noneControl(item['balanced'])))
                
                

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False
    
    
    
    
    def __getMayaProduct(self,po):
        for item in self.mayaPaymentProducts:
            if(item.SiparisNo) == po:
                return self.__noneControl(item.Total)
    def __getMayaPaid(self,po):
        for item in self.mayaPaymentList:
            if(item.SiparisNo) == po:
                return self.__noneControl(item.Total)
    
    
    def __getOrder(self,customer_id):
        for item in self.orders:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Navlun) + self.__noneControl(item.Detay1) + self.__noneControl(item.Detay2) + self.__noneControl(item.Detay3) + self.__noneControl(item.Sigorta)
    
    def __getProduct(self,customer_id):
        for item in self.products:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.SatisToplam)
    
    def __getPaid(self,customer_id):
        for item in self.paids:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Tutar)
            
    
    def __getPaidFilter(self,customer_id):
        for item in self.paidsFilter:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Tutar)
    
    
    def __getOrderForwarding(self,customer_id):
        for item in self.ordersForwarding:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Navlun) + self.__noneControl(item.Detay1) + self.__noneControl(item.Detay2) + self.__noneControl(item.Detay3) + self.__noneControl(item.Sigorta)
    
    def __getOrderProduction(self,customer_id):
        for item in self.ordersProduction:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Navlun) + self.__noneControl(item.Detay1) + self.__noneControl(item.Detay2) + self.__noneControl(item.Detay3) + self.__noneControl(item.Sigorta)
    
    def __getProductForwarding(self,customer_id):
        for item in self.productsForwarding:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.SatisToplam)
            
            
    def __getProductForwardingFilter(self,customer_id):
        for item in self.productsForwardingFilter:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.SatisToplam)
            
            
            
            
    def __getProductProduction(self,customer_id):
        for item in self.productsProduction:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.SatisToplam)    
            
    def __getProductProductionFilter(self,customer_id):
        for item in self.productsProductionFilter:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.SatisToplam)    
    
    def __getPaidForwarding(self,customer_id):
        for item in self.paidsForwarding:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Tutar)
    def __getPaidProduction(self,customer_id):
        for item in self.paidsProduction:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Tutar)
    
    def __getAdvancePayment(self,customer_id):
        for item in self.advancePayment:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Pesinat)
    def __getAdvancePaymentFilter(self,customer_id):
        for item in self.advancePaymentFilter:
            if(item.MusteriID == customer_id):
                return self.__noneControl(item.Pesinat)
    
    def __noneControl(self,value):
        if value == None:
            return 0
        else:
            return float(value)
    
    def setPaidSave(self,data):
        try:
            self.sql.update_insert("""
                                  insert into Odemeler_MekmerTB(Tarih,MusteriID,SiparisNo,FinansOdemeTurID,Aciklama,Tutar,Masraf,KullaniciID,FaturaKesimTurID,Kur)
                                  VALUES(?,?,?,?,?,?,?,?,?,?)
                                  """,(data['tarih'],data['musteri_id'],data['siparisno'],1,data['aciklama'],data['tutar'],data['masraf'],data['KullaniciID'],1,data['kur']))
            
            return True
        except Exception as e :
            print('setPaidSave hata',str(e))
            return False
class FinanceTestDetail:
    def __init__(self,customer_id):
        self.sql = SqlConnect().data
        self.orderForwardingDetail = self.sql.getStoreList("""
                                                    select 
                                                        s.MusteriID,
                                                        s.SiparisNo,
                                                        s.NavlunSatis,
                                                        s.DetayTutar_1,
                                                        s.DetayTutar_2,
                                                        s.DetayTutar_3,
                                                        s.sigorta_tutar_satis,
                                                        s.SiparisTarihi,
														s.YuklemeTarihi,
                                                        s.MayaControl

                                                    from SiparislerTB s

                                                    where 
                                                        s.MusteriID = ? and s.SiparisDurumID = 3
                                                    order by
                                                        s.YuklemeTarihi desc
                                                 
                                                 """,(customer_id))
        self.orderProductionDetail = self.sql.getStoreList("""
                                                    select 
                                                        s.MusteriID,
                                                        s.SiparisNo,
                                                        s.NavlunSatis,
                                                        s.DetayTutar_1,
                                                        s.DetayTutar_2,
                                                        s.DetayTutar_3,
                                                        s.sigorta_tutar_satis,
														s.Pesinat,
                                                        s.SiparisTarihi,
														s.YuklemeTarihi,
                                                        s.MayaControl,
                                                        s.SiparisDurumID
                                                    from SiparislerTB s
                                                    where 
                                                        s.MusteriID = ? and s.SiparisDurumID in (1,2)
                                                 """,(customer_id))
        self.productsDetail = self.sql.getStoreList("""
                                                        select 
                                                            su.SiparisNo,
                                                            sum(su.SatisToplam) as SatisToplam,
                                                            s.MusteriID
                                                        from SiparisUrunTB su 
                                                            inner join SiparislerTB s on s.SiparisNo = su.SiparisNo
                                                        where
                                                            s.MusteriID = ?
                                                        group by
                                                            su.SiparisNo,s.MusteriID
                                                    """,(customer_id))
        self.paidDetail = self.sql.getStoreList("""
                                                    select 
                                                        o.SiparisNo,
                                                        sum(o.Tutar) as Tutar
                                                    from OdemelerTB o 
                                                    where
                                                        o.MusteriID=?
                                                    group by
                                                        o.SiparisNo
                                                """,(customer_id))

        self.paidDateDetail = self.sql.getStoreList("""
                                                        select 
                                                        o.Tarih,
														o.SiparisNo,
														o.Tutar
                                                    from OdemelerTB o
                                                    where 
                                                        o.MusteriID=? and o.SiparisNo in 
                                                        (
															select s.SiparisNo from SiparislerTB s where s.SiparisNo = o.SiparisNo
														)

                                                    order by
                                                        o.Tarih desc
                                                    """,(customer_id))
        
        
        self.byDatePaids = self.sql.getStoreList("""
                                                    select 
                                                        sum(o.Tutar) Tutar,
                                                        o.Tarih
                                                    from OdemelerTB o
                                                    where 
                                                        o.MusteriID=? and o.SiparisNo in (
																							select s.SiparisNo from SiparislerTB s where s.SiparisNo = o.SiparisNo
																						  )
                                                    group by 
                                                        o.Tarih 
                                                    order by
                                                        o.Tarih desc
                                                 """,(customer_id))
    
    def getDetailList(self):
        liste = list()
        for item in self.orderProductionDetail:
            model = FinanceDetailModel()
            model.customer_id = item.MusteriID
            model.po = item.SiparisNo
            model.cost = (
                            self.__noneControl(item.NavlunSatis) + 
                            self.__noneControl(item.DetayTutar_1) + 
                            self.__noneControl(item.DetayTutar_2) + 
                            self.__noneControl(item.DetayTutar_3) +
                            self.__noneControl(item.sigorta_tutar_satis) +
                            self.__noneControl(self.__getProduct(item.SiparisNo))   
                        )
            model.paid = self.__noneControl(self.__getPaid(item.SiparisNo))
            model.balance = (model.cost - model.paid)
            if(item.SiparisDurumID == 2):
                model.status = 'Üretim'
            elif (item.SiparisDurumID == 1):
                model.status = 'Bekleyen'
            model.advanced_payment = self.__noneControl(item.Pesinat)
            model.product_date = item.SiparisTarihi
            model.forwarding_date = item.YuklemeTarihi
            model.maya_control = self.__noneBooleanControl(item.MayaControl)
            model.paid_date = self.__getPaidDate(item.SiparisNo)
            liste.append(model)
        for item in self.orderForwardingDetail:
            model = FinanceDetailModel()
            model.customer_id = item.MusteriID
            model.po = item.SiparisNo
            model.cost = (
                            self.__noneControl(item.NavlunSatis) + 
                            self.__noneControl(item.DetayTutar_1) + 
                            self.__noneControl(item.DetayTutar_2) + 
                            self.__noneControl(item.DetayTutar_3) +
                            self.__noneControl(item.sigorta_tutar_satis) +
                            self.__noneControl(self.__getProduct(item.SiparisNo))   
                        )
            model.paid = self.__noneControl(self.__getPaid(item.SiparisNo))
            model.balance = (model.cost - model.paid)
            model.status = 'Sevk'
            model.product_date = item.SiparisTarihi
            model.forwarding_date = item.YuklemeTarihi
            model.maya_control = self.__noneBooleanControl(item.MayaControl)
            model.paid_date = self.__getPaidDate(item.SiparisNo)
            liste.append(model)
            
        
            

        schema = FinanceDetailSchema(many=True)
        return schema.dump(liste)
    
    def getByDatePaids(self):
        liste = list()
        for item in  self.byDatePaids:
            model = ByDatePaidsModel()
            model.date = item.Tarih
            model.paid = item.Tutar
            liste.append(model)
        schema = ByDatePaidsSchema(many=True)
        return schema.dump(liste)
    
    
    
    def __noneBooleanControl(self,value):
        if(value == None):
            return False
        else:
            return value
    
    def __getProduct(self,po):
        for item in self.productsDetail:
            if(item.SiparisNo == po):
                return self.__noneControl(item.SatisToplam)
    
    def __getPaid(self,po):
        for item in self.paidDetail:
            if(item.SiparisNo == po):
                return self.__noneControl(item.Tutar)
            
    def __getPaidDate(self,po):
        liste = list()
        for item in self.paidDateDetail:
            if(item.SiparisNo == po):
                liste.append({'date':str(item.Tarih),'paid':item.Tutar})
        return liste
    

    
    
    def __noneControl(self,value):
        if(value == None):
            return 0
        else:
            return float(value)
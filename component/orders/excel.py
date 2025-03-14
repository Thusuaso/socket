from flask import jsonify,request
from flask import jsonify,request,send_file
from api.sql import *
from openpyxl import *
from openpyxl.styles import PatternFill,Border,Side
from openpyxl.cell import Cell
from openpyxl.styles import Alignment  

from openpyxl.drawing.image import Image
import datetime 

import shutil
from flask_restful import Resource
from operator import itemgetter, attrgetter
class ExcelCiktiIslem:
    def __init__(self):
        self.sql = SqlConnect().data
    def uretimCikti(self,data_list):  
        try:
            source_path = 'excel/sablonlar/Uretim_list.xlsx'
            target_path = 'excel/dosyalar/Uretim_list.xlsx'
            shutil.copy2(source_path, target_path)
          
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')
             
            satir = 2
            item = 0 
            j = len(data_list)-1
            
            m = 0
            a = 0
            k = 0

            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            while item <= j  :
               
                 
                 sayfa.cell(satir,column=2,value=data_list[item]['OperasyonAdi']).alignment = Alignment(horizontal='center', vertical='center') 
                
                 sayfa.cell(satir,column=4,value=data_list[item]['SiparisTarihi']).alignment = Alignment(horizontal='center', vertical='center') 
                 sayfa.cell(satir,column=6,value=data_list[item]['FirmaAdi']).alignment = Alignment(horizontal='center', vertical='center')
                 sayfa.cell(satir,column=8,value=data_list[item]['SiparisNo']).alignment = Alignment(horizontal='center', vertical='center')    
                 k = satir
                 satir += 1
                 i = 0
               
                 m = 0
                 a=0
                 
                 for item1 in data_list:
                   
                    if data_list[i]['SiparisNo'] == data_list[item]['SiparisNo']:
                         
                         sayfa.cell(satir-1,column=1,value=item1['OperasyonAdi'])
                        
                         sayfa.cell(satir-1,column=3,value=item1['SiparisTarihi'])
                         sayfa.cell(satir-1,column=5,value=item1['FirmaAdi'])
                         sayfa.cell(satir-1,column=7,value=item1['SiparisNo'])
                         sayfa.cell(satir-1,column=9,value=item1['UrunAdi'])
                         sayfa.cell(satir-1,column=10,value=item1['UrunUretimAciklama'])
                         sayfa.cell(satir-1,column=11,value=item1['En'] + 'x'+ item1['Boy']+ 'x'+ item1['Kenar'])
                         sayfa.cell(satir-1,column=12,value=item1['UrunFirmaAdi'])
                         sayfa.cell(satir-1,column=13,value=item1['Miktar'])
                         sayfa.cell(satir-1,column=14,value=item1['Uretim'])
                         if(item1['UrunBirimID'] == 1):
                            sayfa.cell(satir-1,column=15,value='M2')
                         elif(item1['UrunBirimID'] == 2):
                            sayfa.cell(satir-1,column=15,value='Adet')
                         elif(item1['UrunBirimID'] == 3):
                            sayfa.cell(satir-1,column=15,value='MT')
                            
                         remainder = sayfa.cell(satir-1,column=16,value=float(item1['Miktar'] - float(item1['Uretim'])))
                         remainder.border = Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'))
                         satir += 1
                         a += 1
                      
                         i +=1
                         m +=1
                        
                        
                    else :  i +=1
                 satir = satir-2        
                 sayfa.merge_cells(start_row=k,start_column=2,end_row=satir,end_column=2)##.fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)     
                 sayfa.merge_cells(start_row=k,start_column=4,end_row=satir,end_column=4)##.fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)     
                 sayfa.merge_cells(start_row=k,start_column=6,end_row=satir,end_column=6)
                 sayfa.merge_cells(start_row=k,start_column=8,end_row=satir,end_column=8)
                 item = item + m
                 if i <j : i +=1 
                 satir += 1
            

            iscilik = self.__getOzelIscilik()
            item2 = 0 
            j2 = len(iscilik)-1
            
            m2 = 0
            a2 = 0
            k2 = 0
            merge_1 = 'A' + str(satir) + ':' + 'P' + str(satir)
            sayfa.merge_cells(merge_1)
            iscilik_cell = sayfa.cell(satir,column=1,value="ÖZEL İŞÇİLİK YAPILACAKLAR")
            iscilik_cell.border = thin_border
            iscilik_cell.alignment = Alignment(horizontal='center', vertical='center') 
            iscilik_cell.fill = PatternFill(fill_type="solid", start_color='FF' + 'F8312F', end_color='FF' + 'F8312F')

            satir+=1
            while item2 <= j2  :
                 
                 sayfa.cell(satir,column=2,value=iscilik[item2][12]).alignment = Alignment(horizontal='center', vertical='center') 
                
                 sayfa.cell(satir,column=4,value=iscilik[item2][13]).alignment = Alignment(horizontal='center', vertical='center') 
                 sayfa.cell(satir,column=6,value=iscilik[item2][14]).alignment = Alignment(horizontal='center', vertical='center')
                 sayfa.cell(satir,column=8,value=iscilik[item2][1]).alignment = Alignment(horizontal='center', vertical='center')    
                 k2 = satir
                 satir += 1
                 i2 = 0
               
                 m2 = 0
                 a2 = 0



                 for item1 in iscilik:
                   
                    if iscilik[i2][1] == iscilik[item2][1]:
                         
                         sayfa.cell(satir-1,column=1,value=item1[12])
                        
                         sayfa.cell(satir-1,column=3,value=item1[13])
                         sayfa.cell(satir-1,column=5,value=item1[14])
                         sayfa.cell(satir-1,column=7,value=item1[1])
                         sayfa.cell(satir-1,column=9,value=item1[5])
                         sayfa.cell(satir-1,column=10,value=item1[3])
                         sayfa.cell(satir-1,column=11,value=item1[7] + 'x'+ item1[8]+ 'x'+ item1[9])
                         sayfa.cell(satir-1,column=12,value=item1[15])
                         sayfa.cell(satir-1,column=13,value=item1[10])
                         sayfa.cell(satir-1,column=14,value=0)
                         if(item1[16] == 1):
                            sayfa.cell(satir-1,column=15,value='M2')
                         elif(item1[16] == 2):
                            sayfa.cell(satir-1,column=15,value='Adet')
                         elif(item1[16] == 3):
                            sayfa.cell(satir-1,column=15,value='MT')
                            
                         remainder = sayfa.cell(satir-1,column=16,value=0)
                         remainder.border = Border(left=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'))
                         satir += 1
                         a2 += 1
                      
                         i2 +=1
                         m2 +=1
                        
                        
                    else :  i2 +=1
                 satir = satir-2        
                 sayfa.merge_cells(start_row=k2,start_column=2,end_row=satir,end_column=2)##.fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)     
                 sayfa.merge_cells(start_row=k2,start_column=4,end_row=satir,end_column=4)##.fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)     
                 sayfa.merge_cells(start_row=k2,start_column=6,end_row=satir,end_column=6)
                 sayfa.merge_cells(start_row=k2,start_column=8,end_row=satir,end_column=8)
                 item2 = item2 + m2
                 if i2 <j2 : i2 +=1 
                 satir += 1
            


            kitap.save(target_path)
            kitap.close()
            
            return True

        except Exception as e:
            print('ExcelCiktiIslem uretimCikti Hata : ',str(e))
            return False
    def __getOzelIscilik(self):
        
        result = self.sql.getList("""
        select 

	seg.Aciklama IscilikAciklama,
	s.SiparisNo,
	su.MusteriAciklama,
	su.UretimAciklama as UrunUretimAciklama,
	k.KategoriAdi,
	urun.UrunAdi,
	yk.YuzeyIslemAdi,
	ol.En,
	ol.Boy,
	ol.Kenar,
	su.Miktar,
	birim.BirimAdi,
	(select KullaniciAdi from KullaniciTB kullanici where kullanici.ID=s.Operasyon) as OperasyonAdi,
	s.SiparisTarihi,
	(select FirmaAdi from MusterilerTB musteri where musteri.ID = s.MusteriID) as FirmaAdi,
	(select FirmaAdi from TedarikciTB tedarikci where tedarikci.ID = su.TedarikciID) as UrunFirmaAdi,
    su.UrunBirimID


from SiparisEkstraGiderlerTB seg
inner join SiparislerTB s on s.SiparisNo=seg.SiparisNo
inner join SiparisUrunTB su on su.UrunKartID = seg.UrunKartID
inner join UrunKartTB uk on uk.ID = seg.UrunKartID
inner join KategoriTB k on k.ID = uk.KategoriID
inner join UrunlerTB urun on urun.ID = uk.UrunID
inner join YuzeyKenarTB yk on yk.ID = uk.YuzeyID
inner join OlculerTB ol on ol.ID = uk.OlcuID
inner join UrunBirimTB birim on birim.ID = su.UrunBirimID
where s.SiparisDurumID=2


        """)
        if(len(result)>0):
            return result
        else:
            return []


class UretimExcelCiktiApi(Resource):

    def post(self):

        data_list = request.get_json()

        islem = ExcelCiktiIslem()
     
        result = islem.uretimCikti(data_list)
        
        return jsonify({'status' : result})

    def get(self):

        excel_path = 'excel/dosyalar/Uretim_list.xlsx'

        return send_file(excel_path,as_attachment=True)



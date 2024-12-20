from openpyxl import *
import shutil
import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import shutil

class ExcelCiktiIslem:
    def __noneType(self,value):
        if(value == None):
            return 0
        else:
            return value
        
    def maliyet_rapor_ciktisi(self,data_list):
            try:
                source_path = 'excel/sablonlar/ayo_maliyet_listesi.xlsx'
                target_path = 'excel/dosyalar/ayo_maliyet_listesi.xlsx'

                shutil.copy2(source_path, target_path)

               
                kitap = load_workbook(target_path)
                sayfa = kitap.get_sheet_by_name('Sheet')

                satir = 2

                for item in data_list:
                    sayfa.cell(satir,column=1,value=item['siparisci'])
                    sayfa.cell(satir,column=2,value=item['operasyon'])
                    sayfa.cell(satir,column=3,value=item['siparis_no'])
                    sayfa.cell(satir,column=4,value=item['marketing'])
                    sayfa.cell(satir,column=5,value=item['faturatur'])
                    sayfa.cell(satir,column=6,value=item['siparis_tarihi'])             
                    sayfa.cell(satir,column=7,value=item['yukleme_tarihi'])
                    sayfa.cell(satir,column=8,value=item['musteri_adi'])
                    sayfa.cell(satir,column=9,value=item['ulke_adi'])
                    sayfa.cell(satir,column=10,value=item['teslim_sekli'])
                    sayfa.cell(satir,column=11,value=item['toplam_bedel'])
                    sayfa.cell(satir,column=12,value=item['mekmar_alim'])
                    sayfa.cell(satir,column=13,value=item['mekmoz_alim'])
                    sayfa.cell(satir,column=14,value=item['dis_alim'])
                    sayfa.cell(satir,column=15,value=item['nakliye'])
                    sayfa.cell(satir,column=16,value=item['gumruk'])
                    sayfa.cell(satir,column=17,value=item['ilaclama'])
                    sayfa.cell(satir,column=18,value=item['liman'])
                    sayfa.cell(satir,column=19,value=item['sigorta'])
                    sayfa.cell(satir,column=20,value=item['navlun'])
                    sayfa.cell(satir,column=21,value=item['detay_1'])
                    sayfa.cell(satir,column=22,value=item['detay_2'])
                    sayfa.cell(satir,column=23,value=item['detay_3'])
                    sayfa.cell(satir,column=24,value=item['mekus_masraf'])
                   
                   
                    sayfa.cell(satir,column=25,value=item['pazarlama'])
                    sayfa.cell(satir,column=26,value=item['ozel_iscilik'])
                    sayfa.cell(satir,column=27,value=item['banka_masrafi'])
                    sayfa.cell(satir,column=28,value=item['kurye_masrafi'])
                    sayfa.cell(satir,column=29,value=item['masraf_toplam'])
                    sayfa.cell(satir,column=30,value=item['kar_zarar'])
                    sayfa.cell(satir,column=31,value=item['kar_zarar_tl'])
                
                    sayfa.cell(satir,column=32,value=item['dosya_kapanma_date'])
                  
                    
                    
                    

                    satir += 1

                kitap.save(target_path)
                kitap.close()

                return True

            except Exception as e:
                print('ExcelCiktiIslem depoCikti Hata : ',str(e))
                return False
            
    def getMkRaporlariExcelList(self,data):
        try:
            print(data)
            source_path = 'excel/sablonlar/mkRaporlari.xlsx'
            target_path = 'excel/dosyalar/mkRaporlari.xlsx'
            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            
            sayfa.cell(1,column=1,value=datetime.datetime.now().strftime('%Y') + ' YILI BAŞINDAN İTİBAREN ALINAN SİPARİŞLER')
            # sayfa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            satir = 3
            
            byCustomerTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byPo']:
                sayfa.cell(satir,column=1,value=item['tarih'])
                sayfa.cell(satir,column=2,value=item['firma'])
                sayfa.cell(satir,column=3,value=item['po'])
                sayfa.cell(satir,column=4,value=item['teslim'])
                sayfa.cell(satir,column=5,value=item['fob'])
                sayfa.cell(satir,column=6,value=item['ddp'])
                byCustomerTotal['fob'] += item['fob']
                byCustomerTotal['ddp'] += item['ddp']
                satir += 1
            sayfa.cell(satir,column=1,value='Toplam')
            sayfa.cell(satir,column=5,value=byCustomerTotal['fob'])
            sayfa.cell(satir,column=6,value=byCustomerTotal['ddp'])

            sayfa2 = kitap.get_sheet_by_name('Sayfa2')
            sayfa2.cell(1,column=1,value=datetime.datetime.now().strftime('%Y') + ' TARİHİ İTİBARİYLE SİPARİŞLER')
            # sayfa2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            satir2 = 3
            byMarketingTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byMarketing']:
                sayfa2.cell(satir2,column=1,value=item['marketing'])
                sayfa2.cell(satir2,column=2,value=item['toplam'])
                sayfa2.cell(satir2,column=3,value=item['toplamCfr'])
                byMarketingTotal['fob'] += item['toplam']
                byMarketingTotal['ddp'] += item['toplamCfr']
                satir2 += 1
            sayfa2.cell(satir2,column=1,value='Toplam')
            sayfa2.cell(satir2,column=2,value=byMarketingTotal['fob'])
            sayfa2.cell(satir2,column=3,value=byMarketingTotal['ddp'])
            
            sayfa2.cell(1,column=5,value=datetime.datetime.now().strftime('%x') + ' TARİHİ İTİBARİYLE SİPARİŞLER DETAY')
            # sayfa2.merge_cells(start_row=satir2,start_column=1,end_row=satir2,end_column=5)
            satir3=3
            byCustomerTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byCustomer']:
                sayfa2.cell(satir3,column=5,value=item['musteriAdi'])
                sayfa2.cell(satir3,column=6,value=item['marketing'])
                sayfa2.cell(satir3,column=7,value=item['ulkeAdi'])
                sayfa2.cell(satir3,column=8,value=item['toplam'])
                sayfa2.cell(satir3,column=9,value=item['toplamCfr'])
                byCustomerTotal['fob'] += item['toplam']
                byCustomerTotal['ddp'] += item['toplamCfr']
                satir3 += 1
            sayfa2.cell(satir3,column=5,value='Toplam')
            sayfa2.cell(satir3,column=8,value=byCustomerTotal['fob'])
            sayfa2.cell(satir3,column=9,value=byCustomerTotal['ddp'])
            mekmarList = []
            mekmerList = []
            icPiyasaList = []
            imperialHomesList = []

            def siparislerMarketing():
                for item in data['byCustomer']:
                    if(item['marketing'] == 'Mekmar'):
                        mekmarList.append(item)
                    elif item['marketing'] == 'Mekmer':
                        mekmerList.append(item)
                    elif item['marketing'] == 'İç Piyasa':
                        icPiyasaList.append(item)
                    elif item['marketing'] == 'Imperial Homes':
                        imperialHomesList.append(item)
        
            siparislerMarketing()
            
            sayfa2.cell(1,column=11,value='Mekmar')
            satir4 = 3
            mekmarListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in mekmarList:
                sayfa2.cell(satir4,column=11,value=item['musteriAdi'])
                sayfa2.cell(satir4,column=12,value=item['ulkeAdi'])
                sayfa2.cell(satir4,column=13,value=item['toplam'])
                sayfa2.cell(satir4,column=14,value=item['toplamCfr'])
                mekmarListTotal['fob'] += item['toplam']
                mekmarListTotal['ddp'] += item['toplamCfr']
                satir4 += 1
            sayfa2.cell(satir4,column=11,value='Toplam')
            sayfa2.cell(satir4,column=13,value=mekmarListTotal['fob'])
            sayfa2.cell(satir4,column=14,value=mekmarListTotal['ddp'])
            
            sayfa2.cell(1,column=16,value='İç Piyasa')
            satir5 = 3
            icPiyasaListTotal = {
                'fob':0,
                'ddp':0,    
            }
            for item in icPiyasaList:
                sayfa2.cell(satir5,column=16,value=item['musteriAdi'])
                sayfa2.cell(satir5,column=17,value=item['ulkeAdi'])
                sayfa2.cell(satir5,column=18,value=item['toplam'])
                sayfa2.cell(satir5,column=19,value=item['toplamCfr'])
                icPiyasaListTotal['fob'] += item['toplam']
                icPiyasaListTotal['ddp'] += item['toplamCfr']
                satir5 += 1
            sayfa2.cell(satir5,column=16,value='Toplam')
            sayfa2.cell(satir5,column=18,value=icPiyasaListTotal['fob'])
            sayfa2.cell(satir5,column=19,value=icPiyasaListTotal['ddp'])
            
            sayfa2.cell(1,column=21,value='Mekmer')
            satir6 = 3
            mekmerTotal = {
                'fob':0,
                'ddp':0,    
            }
            for item in mekmerList:
                sayfa2.cell(satir6,column=21,value=item['musteriAdi'])
                sayfa2.cell(satir6,column=22,value=item['ulkeAdi'])
                sayfa2.cell(satir6,column=23,value=item['toplam'])
                sayfa2.cell(satir6,column=24,value=item['toplamCfr'])
                mekmerTotal['fob'] += item['toplam']
                mekmerTotal['ddp'] += item['toplamCfr']
                satir6 += 1
            sayfa2.cell(satir6,column=21,value='Toplam')
            sayfa2.cell(satir6,column=23,value=mekmerTotal['fob'])
            sayfa2.cell(satir6,column=24,value=mekmerTotal['ddp'])
            
            sayfa2.cell(1,column=26,value='İmperial Homes')
            satir7 = 3
            imperialHomesTotal = {
                'fob':0,
                'ddp':0,    
            }
            for item in imperialHomesList:
                sayfa2.cell(satir7,column=26,value=item['musteriAdi'])
                sayfa2.cell(satir7,column=27,value=item['ulkeAdi'])
                sayfa2.cell(satir7,column=28,value=item['toplam'])
                sayfa2.cell(satir7,column=29,value=item['toplamCfr'])
                imperialHomesTotal['fob'] += item['toplam']
                imperialHomesTotal['ddp'] += item['toplamCfr']
                satir7 += 1
            sayfa2.cell(satir7,column=26,value='Toplam')
            sayfa2.cell(satir7,column=28,value=imperialHomesTotal['fob'])
            sayfa2.cell(satir7,column=29,value=imperialHomesTotal['ddp'])

            sayfa3 = kitap.get_sheet_by_name('Sayfa3')
            sayfa3.cell(1,column=1,value=  '1-1-2024' + ' ' + datetime.datetime.now().strftime('%x') + ' ARASI YÜKLEMELER')
            satir8 = 3
            byMarketingYuklemeTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in data['byMarketingYukleme']:
                sayfa3.cell(satir8,column=1,value=item['marketing'])
                sayfa3.cell(satir8,column=2,value=item['fobToplam'])
                sayfa3.cell(satir8,column=3,value=item['cfrToplam'])
                byMarketingYuklemeTotal['fob'] += item['fobToplam']
                byMarketingYuklemeTotal['ddp'] += item['cfrToplam']
                satir8 += 1
            sayfa3.cell(satir8,column=1,value='Toplam')
            sayfa3.cell(satir8,column=2,value=byMarketingYuklemeTotal['fob'])
            sayfa3.cell(satir8,column=3,value=byMarketingYuklemeTotal['ddp'])
            
            sayfa4 = kitap.get_sheet_by_name('Sayfa4')
            satir13 = 2
            for item in data['byCustomerOrder']:
                sayfa4.cell(satir13,column=1,value=item['musteri'])
                sayfa4.cell(satir13,column=2,value=item['ulkeAdi'])
                sayfa4.cell(satir13,column=3,value=item['temsilci'])
                
                if item['Toplam'] == None:
                    sayfa4.cell(satir13,column=4,value=0)
                else:
                    sayfa4.cell(satir13,column=4,value=item['Toplam'])
                
                
                
                if item['BuYilUretim'] == None:
                    sayfa4.cell(satir13,column=5,value=0) 
                else:
                    sayfa4.cell(satir13,column=5,value=item['BuYilUretim']) 

                if item['BuYilSevkiyat'] == None:
                    sayfa4.cell(satir13,column=6,value=0) 
                else:
                    sayfa4.cell(satir13,column=6,value=item['BuYilSevkiyat']) 
                
                

                if item['GecenYil'] == None:
                    sayfa4.cell(satir13,column=7,value=0) 
                else:
                    sayfa4.cell(satir13,column=7,value=item['GecenYil']) 

                if item['OncekiYil'] == None:
                    sayfa4.cell(satir13,column=8,value=0) 
                else:
                    sayfa4.cell(satir13,column=8,value=item['OncekiYil'])
                    
                if item['OnDokuzYili'] == None:
                    sayfa4.cell(satir13,column=9,value=0) 
                else:
                    sayfa4.cell(satir13,column=9,value=item['OnDokuzYili']) 
                    
                if item['OnSekizYili'] == None:
                    sayfa4.cell(satir13,column=10,value=0) 
                else:
                    sayfa4.cell(satir13,column=10,value=item['OnSekizYili'])                 
                
                if item['OnYediYili'] == None:
                    sayfa4.cell(satir13,column=11,value=0) 
                else:
                    sayfa4.cell(satir13,column=11,value=item['OnYediYili']) 
                
                if item['OnAltiYili'] == None:
                    sayfa4.cell(satir13,column=12,value=0) 
                else:
                    sayfa4.cell(satir13,column=12,value=item['OnAltiYili']) 
                    
                if item['OnBesYili'] == None:
                    sayfa4.cell(satir13,column=13,value=0) 
                else:
                    sayfa4.cell(satir13,column=13,value=item['OnBesYili']) 
                
                
                if item['OnDortYili'] == None:
                    sayfa4.cell(satir13,column=14,value=0) 
                else:
                    sayfa4.cell(satir13,column=14,value=item['OnDortYili']) 
                
                if item['OnUcYili'] == None:
                    sayfa4.cell(satir13,column=15,value=0) 
                else:
                    sayfa4.cell(satir13,column=15,value=item['OnUcYili']) 
                
                if item['OnUcYiliOncesi'] == None:
                    sayfa4.cell(satir13,column=16,value=0) 
                else:
                    sayfa4.cell(satir13,column=16,value=item['OnUcYiliOncesi'])
                satir13 += 1
            
            
            mekmarYuklemeList = []
            mekmerYuklemeList = []
            icPiyasaYuklemeList = []
            imperialHomesYuklemeList = []

            def yuklemelerMarketing():
                for item in data['byMarketingDetayYukleme']:
                    if(item['marketing'] == 'Mekmar'):
                        mekmarYuklemeList.append(item)
                    elif item['marketing'] == 'Mekmer':
                        mekmerYuklemeList.append(item)
                    elif item['marketing'] == 'İç Piyasa':
                        icPiyasaYuklemeList.append(item)
                    elif item['marketing'] == 'Imperial Homes':
                        imperialHomesYuklemeList.append(item)
            yuklemelerMarketing()
            
            sayfa3.cell(1,column=5,value=  'Mekmar')
            satir9 = 3
            mekmarYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in mekmarYuklemeList:
                sayfa3.cell(satir9,column=5,value=item['musteri'])
                sayfa3.cell(satir9,column=6,value=item['toplamFob'])
                sayfa3.cell(satir9,column=7,value=item['toplamCfr'])
                mekmarYuklemeListTotal['fob'] += item['toplamFob']
                mekmarYuklemeListTotal['ddp'] += item['toplamCfr']
                satir9 += 1
            sayfa3.cell(satir9,column=5,value='Toplam')
            sayfa3.cell(satir9,column=6,value=mekmarYuklemeListTotal['fob'])
            sayfa3.cell(satir9,column=7,value=mekmarYuklemeListTotal['ddp'])
            
            sayfa3.cell(1,column=9,value=  'İç Piyasa')
            satir10 = 3
            icpiyasaYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in icPiyasaYuklemeList:
                sayfa3.cell(satir10,column=9,value=item['musteri'])
                sayfa3.cell(satir10,column=10,value=item['toplamFob'])
                sayfa3.cell(satir10,column=11,value=item['toplamCfr'])
                icpiyasaYuklemeListTotal['fob'] += item['toplamFob']
                icpiyasaYuklemeListTotal['ddp'] += item['toplamCfr']
                satir10 += 1
            sayfa3.cell(satir10,column=9,value='Toplam')
            sayfa3.cell(satir10,column=10,value=icpiyasaYuklemeListTotal['fob'])
            sayfa3.cell(satir10,column=11,value=icpiyasaYuklemeListTotal['ddp'])
            
            sayfa3.cell(1,column=13,value=  'Mekmer')
            satir11 = 3
            mekmerYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in mekmerYuklemeList:
                sayfa3.cell(satir11,column=13,value=item['musteri'])
                sayfa3.cell(satir11,column=14,value=item['toplamFob'])
                sayfa3.cell(satir11,column=15,value=item['toplamCfr'])
                mekmerYuklemeListTotal['fob'] += item['toplamFob']
                mekmerYuklemeListTotal['ddp'] += item['toplamCfr']
                satir11 += 1
            sayfa3.cell(satir11,column=13,value='Toplam')
            sayfa3.cell(satir11,column=14,value=mekmerYuklemeListTotal['fob'])
            sayfa3.cell(satir11,column=15,value=mekmerYuklemeListTotal['ddp'])
            
            sayfa3.cell(1,column=17,value=  'İmperial Homes')
            satir12 = 3
            imperialHomesYuklemeListTotal = {
                'fob':0,
                'ddp':0,
            }
            for item in imperialHomesYuklemeList:
                sayfa3.cell(satir12,column=17,value=item['musteri'])
                sayfa3.cell(satir12,column=18,value=item['toplamFob'])
                sayfa3.cell(satir12,column=19,value=item['toplamCfr'])
                imperialHomesYuklemeListTotal['fob'] += item['toplamFob']
                imperialHomesYuklemeListTotal['ddp'] += item['toplamCfr']
                satir12 += 1
            sayfa3.cell(satir12,column=17,value='Toplam')
            sayfa3.cell(satir12,column=18,value=imperialHomesYuklemeListTotal['fob'])
            sayfa3.cell(satir12,column=19,value=imperialHomesYuklemeListTotal['ddp'])
            
            sayfa5 = kitap.get_sheet_by_name('Sayfa5')
            satir15=2
            for item in data['byYuklemevSiparisler']:
                sayfa5.cell(satir15,column=1,value=item['musteriadi'])
                sayfa5.cell(satir15,column=2,value=item['siparisfob'])
                sayfa5.cell(satir15,column=3,value=item['yuklenenddp'])
                satir15 += 1
            

                
            
            
            kitap.save(target_path)
            kitap.close()
            return True

            
            
        except Exception as e:
            print('getMkRaporlariExcelList hata',str(e))
            return False
        
    def stok_rapor_ciktisi(self,data_list):
        try:
            source_path = 'excel/sablonlar/Stok_listesi.xlsx'
            target_path = 'excel/dosyalar/Stok_listesi.xlsx'

            shutil.copy2(source_path, target_path)
            rgb=[204,102,0]
            color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
            border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            kalin = Font(bold=True,size=16)
            satir = 1
            sayfa.cell(satir,column=1,value='Category').fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=1,value='Category').font = kalin
            sayfa.cell(satir,column=2,value= "Selection" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=2,value= "Selection" ).font = kalin
            sayfa.cell(satir,column=3,value= "Surface" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=3,value= "Surface" ).font = kalin


            sayfa.cell(satir,column=4,value='W').fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=4,value='W').font = kalin
            sayfa.cell(satir,column=5,value='L').fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=5,value='L').font = kalin
            sayfa.cell(satir,column=6,value='Thickness').fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=6,value='Thickness').font = kalin
            sayfa.cell(satir,column=7,value= "M2" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=7,value= "M2" ).font = kalin
            sayfa.cell(satir,column=8,value= "Crate Amount" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=8,value= "Crate Amount" ).font = kalin


            
            
            satir += 1
            for item1 in data_list:
                cell1 = sayfa.cell(satir,column=1,value=item1['KategoriAdi'])
                cell1.border = border
                cell6 = sayfa.cell(satir,column=2,value=item1['UrunAdi'])
                cell6.border = border
                cell5 = sayfa.cell(satir,column=3,value=item1['YuzeyIslemAdi'])
                cell5.border = border
                cell2 = sayfa.cell(satir,column=4,value=item1['En'])
                cell2.border = border
                cell3 = sayfa.cell(satir,column=5,value=item1['Boy'])
                cell3.border = border
                cell4 = sayfa.cell(satir,column=6,value=item1['Kenar'])
                cell4.border = border

                cell8 = sayfa.cell(satir,column=7,value=item1['Toplam'])
                cell8.border = border
                cell7 = sayfa.cell(satir,column=8,value=item1['KasaSayisi'])
                cell7.border = border

                satir += 1
            

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False   
    def __months(self,val):
        months = {
            1:'January',
            2:'February',
            3:'March',
            4:'April',
            5:'May',
            6:'June',
            7:'July',
            8:'August',
            9:'September',
            10:'October',
            11:'November',
            12:'December'
        }
        return months[val]
        
    def getGuRaporlariSatisciOperasyonOrders(self,data):
        try:
                source_path = 'excel/sablonlar/gu_reports_summary.xlsx'
                target_path = 'excel/dosyalar/gu_reports_summary.xlsx'
                date = datetime.datetime.now()

                shutil.copy2(source_path, target_path)
                kitap = load_workbook(target_path)
                sayfa = kitap.get_sheet_by_name('Sayfa1')

                satir = 3
                thisYearTotalFob = 0
                thisYearTotalDdp = 0
                sayfa.cell(row=1,column=1,value=str(date.year) + ' Seller Orders')
                for item in data['thisYearSeller']:
                    sayfa.cell(satir,column=1,value=self.__months(item['Month']))
                    sayfa.cell(satir,column=2,value=item['FOB'])
                    sayfa.cell(satir,column=3,value=item['DDP'])
                    thisYearTotalFob += item['FOB']
                    thisYearTotalDdp += item['DDP']
                    satir += 1
                    
                sayfa.cell(row=14,column=2,value=thisYearTotalFob)
                sayfa.cell(row=14,column=3,value=thisYearTotalDdp)
                
                
                previousYearTotalFob = 0
                previousYearTotalDdp = 0
                satir2 = 3
                sayfa.cell(row=1,column=5,value=str(date.year - 1) + ' Seller Orders')
                for item in data['previousYearSeller']:
                    sayfa.cell(satir2,column=5,value=self.__months(item['Month']))
                    sayfa.cell(satir2,column=6,value=item['FOB'])
                    sayfa.cell(satir2,column=7,value=item['DDP'])
                    previousYearTotalFob += item['FOB']
                    previousYearTotalDdp += item['DDP']
                    satir2 += 1
                    
                sayfa.cell(row=14,column=6,value=previousYearTotalFob)
                sayfa.cell(row=14,column=7,value=previousYearTotalDdp)
                
                
                twoYearTotalFob = 0
                twoYearTotalDdp = 0
                satir3 = 3
                sayfa.cell(row=1,column=9,value=str(date.year - 2) + ' Seller Orders')
                for item in data['twoYearAgoYearSeller']:
                    sayfa.cell(satir3,column=9,value=self.__months(item['Month']))
                    sayfa.cell(satir3,column=10,value=item['FOB'])
                    sayfa.cell(satir3,column=11,value=item['DDP'])
                    twoYearTotalFob += item['FOB']
                    twoYearTotalDdp += item['DDP']
                    satir3 += 1
                    
                sayfa.cell(row=14,column=10,value=twoYearTotalFob)
                sayfa.cell(row=14,column=11,value=twoYearTotalDdp)
                
                
                
                
                satir4 = 18
                thisYearOpTotalFob = 0
                thisYearOpTotalDdp = 0
                sayfa.cell(row=16,column=1,value=str(date.year) + ' Operation Orders')
                for item in data['thisYearOperation']:
                    sayfa.cell(satir4,column=1,value=self.__months(item['Month']))
                    sayfa.cell(satir4,column=2,value=item['FOB'])
                    sayfa.cell(satir4,column=3,value=item['DDP'])
                    thisYearOpTotalFob += item['FOB']
                    thisYearOpTotalDdp += item['DDP']
                    satir4 += 1
                    
                sayfa.cell(row=29,column=2,value=thisYearOpTotalFob)
                sayfa.cell(row=29,column=3,value=thisYearOpTotalDdp)
                
                
                
                
                
                satir5 = 18
                previousYearOpTotalFob = 0
                previousYearOpTotalDdp = 0
                sayfa.cell(row=16,column=5,value=str(date.year - 1) + ' Operation Orders')
                for item in data['previousYearOperation']:
                    sayfa.cell(satir5,column=5,value=self.__months(item['Month']))
                    sayfa.cell(satir5,column=6,value=item['FOB'])
                    sayfa.cell(satir5,column=7,value=item['DDP'])
                    previousYearOpTotalFob += item['FOB']
                    previousYearOpTotalDdp += item['DDP']
                    satir5 += 1
                    
                sayfa.cell(row=29,column=6,value=previousYearOpTotalFob)
                sayfa.cell(row=29,column=7,value=previousYearOpTotalDdp)
                
                
                
                satir6 = 18
                twoYearOpTotalFob = 0
                twoYearOpTotalDdp = 0
                sayfa.cell(row=16,column=9,value=str(date.year - 2) + ' Operation Orders')
                for item in data['twoYearAgoOperation']:
                    sayfa.cell(satir6,column=9,value=self.__months(item['Month']))
                    sayfa.cell(satir6,column=10,value=item['FOB'])
                    sayfa.cell(satir6,column=11,value=item['DDP'])
                    twoYearOpTotalFob += item['FOB']
                    twoYearOpTotalDdp += item['DDP']
                    satir6 += 1
                    
                sayfa.cell(row=29,column=10,value=twoYearOpTotalFob)
                sayfa.cell(row=29,column=11,value=twoYearOpTotalDdp)
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
                
            
                    



                kitap.save(target_path)
                kitap.close()

                return True
        except Exception as e:
            print('getGuRaporlariSatisciOperasyonOrders hata',str(e))
            return False
        
    def getGuRaporlariSatisciOperasyonShipped(self,data):
        try:
                source_path = 'excel/sablonlar/gu_reports_summary.xlsx'
                target_path = 'excel/dosyalar/gu_reports_summary.xlsx'
                date = datetime.datetime.now()

                shutil.copy2(source_path, target_path)
                kitap = load_workbook(target_path)
                sayfa = kitap.get_sheet_by_name('Sayfa1')

                satir = 3
                thisYearTotalFob = 0
                thisYearTotalDdp = 0
                sayfa.cell(row=1,column=1,value=str(date.year) + ' Seller Shipped')
                for item in data['thisYearSeller']:
                    sayfa.cell(satir,column=1,value=self.__months(item['Month']))
                    sayfa.cell(satir,column=2,value=item['FOB'])
                    sayfa.cell(satir,column=3,value=item['DDP'])
                    thisYearTotalFob += item['FOB']
                    thisYearTotalDdp += item['DDP']
                    satir += 1
                    
                sayfa.cell(row=14,column=2,value=thisYearTotalFob)
                sayfa.cell(row=14,column=3,value=thisYearTotalDdp)
                
                
                previousYearTotalFob = 0
                previousYearTotalDdp = 0
                satir2 = 3
                sayfa.cell(row=1,column=5,value=str(date.year - 1) + ' Seller Shipped')
                for item in data['previousYearSeller']:
                    sayfa.cell(satir2,column=5,value=self.__months(item['Month']))
                    sayfa.cell(satir2,column=6,value=item['FOB'])
                    sayfa.cell(satir2,column=7,value=item['DDP'])
                    previousYearTotalFob += item['FOB']
                    previousYearTotalDdp += item['DDP']
                    satir2 += 1
                    
                sayfa.cell(row=14,column=6,value=previousYearTotalFob)
                sayfa.cell(row=14,column=7,value=previousYearTotalDdp)
                
                
                twoYearTotalFob = 0
                twoYearTotalDdp = 0
                satir3 = 3
                sayfa.cell(row=1,column=9,value=str(date.year - 2) + ' Seller Shipped')
                for item in data['twoYearAgoYearSeller']:
                    sayfa.cell(satir3,column=9,value=self.__months(item['Month']))
                    sayfa.cell(satir3,column=10,value=item['FOB'])
                    sayfa.cell(satir3,column=11,value=item['DDP'])
                    twoYearTotalFob += item['FOB']
                    twoYearTotalDdp += item['DDP']
                    satir3 += 1
                    
                sayfa.cell(row=14,column=10,value=twoYearTotalFob)
                sayfa.cell(row=14,column=11,value=twoYearTotalDdp)
                
                
                
                
                satir4 = 18
                thisYearOpTotalFob = 0
                thisYearOpTotalDdp = 0
                sayfa.cell(row=16,column=1,value=str(date.year) + ' Operation Shipped')
                for item in data['thisYearOperation']:
                    sayfa.cell(satir4,column=1,value=self.__months(item['Month']))
                    sayfa.cell(satir4,column=2,value=item['FOB'])
                    sayfa.cell(satir4,column=3,value=item['DDP'])
                    thisYearOpTotalFob += item['FOB']
                    thisYearOpTotalDdp += item['DDP']
                    satir4 += 1
                    
                sayfa.cell(row=29,column=2,value=thisYearOpTotalFob)
                sayfa.cell(row=29,column=3,value=thisYearOpTotalDdp)
                
                
                
                
                
                satir5 = 18
                previousYearOpTotalFob = 0
                previousYearOpTotalDdp = 0
                sayfa.cell(row=16,column=5,value=str(date.year - 1) + ' Operation Shipped')
                for item in data['previousYearOperation']:
                    sayfa.cell(satir5,column=5,value=self.__months(item['Month']))
                    sayfa.cell(satir5,column=6,value=item['FOB'])
                    sayfa.cell(satir5,column=7,value=item['DDP'])
                    previousYearOpTotalFob += item['FOB']
                    previousYearOpTotalDdp += item['DDP']
                    satir5 += 1
                    
                sayfa.cell(row=29,column=6,value=previousYearOpTotalFob)
                sayfa.cell(row=29,column=7,value=previousYearOpTotalDdp)
                
                
                
                satir6 = 18
                twoYearOpTotalFob = 0
                twoYearOpTotalDdp = 0
                sayfa.cell(row=16,column=9,value=str(date.year - 2) + ' Operation Shipped')
                for item in data['twoYearAgoOperation']:
                    sayfa.cell(satir6,column=9,value=self.__months(item['Month']))
                    sayfa.cell(satir6,column=10,value=item['FOB'])
                    sayfa.cell(satir6,column=11,value=item['DDP'])
                    twoYearOpTotalFob += item['FOB']
                    twoYearOpTotalDdp += item['DDP']
                    satir6 += 1
                    
                sayfa.cell(row=29,column=10,value=twoYearOpTotalFob)
                sayfa.cell(row=29,column=11,value=twoYearOpTotalDdp)
                

                kitap.save(target_path)
                kitap.close()

                return True
        except Exception as e:
            print('getGuRaporlariSatisciOperasyonOrders hata',str(e))
            return False
        
    def getAyoCostExcel(self,data):
        try:
            source_path = 'excel/sablonlar/ayo_cost_excel.xlsx'
            target_path = 'excel/dosyalar/ayo_cost_excel.xlsx'

            shutil.copy2(source_path, target_path)

            
            kitap = load_workbook(target_path)
            creditcard = kitap.get_sheet_by_name('creditcard')
            travel = kitap.get_sheet_by_name('travel')
            wage = kitap.get_sheet_by_name('wage')
            sample = kitap.get_sheet_by_name('sample')
            other = kitap.get_sheet_by_name('other')


            credit_satir = 2
            credit_tl = 0
            credit_usd = 0
            for item in data['credit']:
                creditcard.cell(credit_satir,column=1,value=item['month'])
                creditcard.cell(credit_satir,column=2,value=item['value'])
                creditcard.cell(credit_satir,column=3,value=item['usd'])
                creditcard.cell(credit_satir,column=4,value=item['currency'])
                credit_tl += item['value']
                credit_usd += item['usd']


                credit_satir += 1
            creditcard.cell(credit_satir,column=2,value = credit_tl)
            creditcard.cell(credit_satir,column=3,value = credit_usd)


            
            travel_satir = 2
            travel_tl = 0
            travel_usd = 0
            for item in data['travel']:
                travel.cell(travel_satir,column=1,value=item['month'])
                travel.cell(travel_satir,column=2,value=item['value'])
                travel.cell(travel_satir,column=3,value=item['usd'])
                travel.cell(travel_satir,column=4,value=item['currency'])
                travel_tl += item['value']
                travel_usd += item['usd']
                travel_satir += 1
            travel.cell(travel_satir,column=2,value = travel_tl)
            travel.cell(travel_satir,column=3,value = travel_usd)


            wage_satir = 2
            wage_tl = 0
            wage_usd = 0
            for item in data['wage']:
                wage.cell(wage_satir,column=1,value=item['month'])
                wage.cell(wage_satir,column=2,value=item['value'])
                wage.cell(wage_satir,column=3,value=item['usd'])
                wage.cell(wage_satir,column=4,value=item['currency'])
                wage_tl += item['value']
                wage_usd += item['usd']
                wage_satir += 1
            wage.cell(wage_satir,column=2,value = wage_tl)
            wage.cell(wage_satir,column=3,value = wage_usd)


            sample_satir = 2
            sample_tl = 0
            sample_usd = 0
            for item in data['sample']:
                sample.cell(sample_satir,column=1,value=item['month'])
                sample.cell(sample_satir,column=2,value=item['value'])
                sample.cell(sample_satir,column=3,value=item['usd'])
                sample.cell(sample_satir,column=4,value=item['currency'])
                sample_tl += item['value']
                sample_usd += item['usd']
                sample_satir += 1
            sample.cell(sample_satir,column=2,value = sample_tl)
            sample.cell(sample_satir,column=3,value = sample_usd)

            other_satir = 2
            other_tl = 0
            other_usd = 0
            for item in data['other']:
                other.cell(other_satir,column=1,value=item['month'])
                other.cell(other_satir,column=2,value=item['value'])
                other.cell(other_satir,column=3,value=item['usd'])
                other.cell(other_satir,column=4,value=item['currency'])
                other_tl += item['value']
                other_usd += item['usd']
                other_satir += 1
            other.cell(other_satir,column=2,value = other_tl)
            other.cell(other_satir,column=3,value = other_usd)
            
            

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False
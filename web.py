import numpy as np
from flask import Flask,jsonify,send_file,request
from flask_restful import Api,Resource
from flask_cors import CORS,cross_origin 
import datetime
app = Flask(__name__)
api = Api(app)
CORS(app, resources={r'/*': {'origins': '*'}})
from openpyxl import *
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font,Alignment,PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import shutil

from component.main import *
from component.finance.finance import *
from component.orders.excel import *
from component.mk.mk import *
from component.stock.stock import *
from component.currency import *

from api.sql import *
import calendar

from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
import smtplib
from email.mime.application import MIMEApplication


class ExcellCiktiIslem:

    def ceki_listesi_excel(self,data_list):
         try:
            source_path = 'excel/sablonlar/ceki_listesi.xlsx'
            target_path = 'excel/dosyalar/ceki_listesi.xlsx'

            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sheet']

            satir = 11
            index = 1
            total_m2 = 0
            total_piece = 0
            total_mt = 0
            total_ton_1 = 0
            total_ton_2 = 0
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            font = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            po = data_list['po']
            sayfa.cell(6,column=1,value= po + ' Packing List')
            for item in data_list['list']:
                sayfa.cell(satir,column=1,value=index).border = thin_border
                sayfa.cell(satir,column=2,value=item['KasaNo']).border = thin_border
                sayfa.cell(satir,column=3,value=item['KategoriAdi']).border = thin_border
                sayfa.cell(satir,column=4,value=item['YuzeyIslem']).border = thin_border
                sayfa.cell(satir,column=5,value=item['UrunAdi']).border = thin_border
                sayfa.cell(satir,column=6,value=item['Kenar']).border = thin_border
                sayfa.cell(satir,column=7,value=item['En']).border = thin_border
                sayfa.cell(satir,column=8,value=item['Boy']).border = thin_border
                sayfa.cell(satir,column=9,value=item['KutuAdet']).border = thin_border
                sayfa.cell(satir,column=10,value=item['Adet']).border = thin_border

                miktar = 0
                kutu = int(item['KutuAdet'])
                if(item['BirimAdi'] == 'Sqm'):
                    if(item['En']=='ANT' and item['Boy']=='PAT'):
                        miktar = float(round((0.74338688 * kutu),2))
                        
                    elif(item['En']=='20,3' and item['Boy']=='SET'):
                        miktar=float(round((0.494914 * kutu),2))
                    elif(item['En'] == 'VAR') or (item['En'] == 'Various') or (item['En'] == '1 LT'):
                        miktar = float(item['Miktar'])
                    else:
                        miktar = float(item['Miktar'])
                elif (item['BirimAdi'] == 'Pcs'):
                    
                    if(float(item['Miktar']) != None or float(item['Miktar']) != 0):
                        
                        miktar = float(item['Miktar'])
                    else:
                        miktar = '-'
                elif (item['BirimAdi'] == 'Mt'):
                    miktar=float(item['Miktar'])

                if(item['BirimAdi'] == 'Sqm'):
                    total_m2 += miktar
                    sayfa.cell(satir,column=11,value=miktar).border = thin_border
                else:
                    sayfa.cell(satir,column=11,value="").border = thin_border

                if(item['BirimAdi'] == 'Pcs'):
                    total_piece += miktar
                    sayfa.cell(satir,column=12,value=miktar).border = thin_border
                else:
                    sayfa.cell(satir,column=12,value="").border = thin_border
                if(item['BirimAdi'] == 'Mt'):
                    total_mt += miktar
                    sayfa.cell(satir,column=13,value=miktar).border = thin_border
                else:
                    sayfa.cell(satir,column=13,value="").border = thin_border

                
                if(item['Ton'] == None or item['Ton'] == 'None' or item['Ton'] == 'undefined'):
                    sayfa.cell(satir,column=14,value=0).border = thin_border
                    sayfa.cell(satir,column=15,value = 0).border = thin_border
                else:
                    total_ton_1 += int(item['Ton'])
                    total_ton_2 += int(item['Ton']) + 30
                    sayfa.cell(satir,column=14,value=int(item['Ton'])).border = thin_border
                    sayfa.cell(satir,column=15,value=int(item['Ton']) + 30).border = thin_border
                
                
                
                satir += 1
                index += 1

            m2font = sayfa.cell(satir,column=11,value=total_m2)
            piecefont =sayfa.cell(satir,column=12,value=total_piece)
            mtfont =sayfa.cell(satir,column=13,value=total_mt)
            ton1font =sayfa.cell(satir,column=14,value=int(total_ton_1))
            ton2font =sayfa.cell(satir,column=15,value=int(total_ton_2))
            m2font.font = font
            piecefont.font = font
            mtfont.font = font
            ton1font.font = font
            ton2font.font = font
            m2font.border = thin_border
            piecefont.border = thin_border
            mtfont.border = thin_border
            ton1font.border = thin_border
            ton2font.border = thin_border

            satir += 2
            merge_1 = 'A' + str(satir) + ':' + 'B' + str(satir)
            sayfa.merge_cells(merge_1)
            total_cases = sayfa.cell(satir,column=1,value="TOTAL CASES")
            sayfa.cell(satir,column=2).border = thin_border
            total_cases.border = thin_border
            total_cases.font = font
            total_cases.alignment  = Alignment(horizontal="center", vertical="center")
            merge_2 = 'C' + str(satir) + ':' + 'D' + str(satir)
            sayfa.merge_cells(merge_2)
            case_total = sayfa.cell(satir,column = 3,value=str(index -1) + ' ' + 'W.CASES')
            sayfa.cell(satir,column=4).border = thin_border
            case_total.border = thin_border
            case_total.font = font
            case_total.alignment  = Alignment(horizontal="center", vertical="center")

            satir += 2

            merge_3 = 'A' + str(satir) + ':' + 'B' + str(satir)
            sayfa.merge_cells(merge_3)
            total_ton = sayfa.cell(satir,column=1,value="TOTAL NET KG")
            sayfa.cell(satir,column=2).border = thin_border
            total_ton.border = thin_border
            total_ton.font = font
            total_ton.alignment  = Alignment(horizontal="center", vertical="center")

            merge_4 = 'C' + str(satir) + ':' + 'D' + str(satir)
            sayfa.merge_cells(merge_4)
            sum_ton = sayfa.cell(satir,column = 3,value=str(total_ton_1) + ' ' + 'kgs')
            sayfa.cell(satir,column=4).border = thin_border
            sum_ton.border = thin_border
            sum_ton.font = font
            sum_ton.alignment  = Alignment(horizontal="center", vertical="center")


            satir += 2

            merge_5 = 'A' + str(satir) + ':' + 'B' + str(satir)
            sayfa.merge_cells(merge_5)
            total_ton_gross = sayfa.cell(satir,column=1,value="TOTAL GROSS KG ")
            sayfa.cell(satir,column=2).border = thin_border
            total_ton_gross.border = thin_border
            total_ton_gross.font = font
            total_ton_gross.alignment  = Alignment(horizontal="center", vertical="center")

            merge_6 = 'C' + str(satir) + ':' + 'D' + str(satir)
            sayfa.merge_cells(merge_6)
            sum_gross_ton = sayfa.cell(satir,column = 3,value=str(total_ton_2) + ' ' + 'kgs')
            sayfa.cell(satir,column=4).border = thin_border
            sum_gross_ton.border = thin_border
            sum_gross_ton.font = font
            sum_gross_ton.alignment  = Alignment(horizontal="center", vertical="center")

            satir += 2

            merge_7 = 'A' + str(satir) + ':' +'D' + str(satir)
            sayfa.merge_cells(merge_7)
            description = sayfa.cell(satir,column = 1,value="EK AÇIKLAMALAR")
            sayfa.cell(satir,column=2).border = thin_border
            sayfa.cell(satir,column=3).border = thin_border
            sayfa.cell(satir,column=4).border = thin_border


            description.border = thin_border
            description.font = font
            description.alignment  = Alignment(horizontal="center", vertical="center")


            merge_8 = 'A' + str(satir + 1) + ':' +'D' + str(satir + 3)
            sayfa.merge_cells(merge_8)
            sayfa.cell(satir + 1,column=1).border = thin_border
            sayfa.cell(satir + 1,column=2).border = thin_border
            sayfa.cell(satir + 1,column=3).border = thin_border
            sayfa.cell(satir + 1,column=4).border = thin_border

            sayfa.cell(satir + 2,column=1).border = thin_border
            sayfa.cell(satir + 2,column=2).border = thin_border
            sayfa.cell(satir + 2,column=3).border = thin_border
            sayfa.cell(satir + 2,column=4).border = thin_border

            sayfa.cell(satir + 3,column=1).border = thin_border
            sayfa.cell(satir + 3,column=2).border = thin_border
            sayfa.cell(satir + 3,column=3).border = thin_border
            sayfa.cell(satir + 3,column=4).border = thin_border





            merge_8 = 'E' + str(satir) + ':' +'O' + str(satir)
            sayfa.merge_cells(merge_8)
            crates_size = sayfa.cell(satir,column = 5,value="KASA ÖLÇÜLERİ")
            sayfa.cell(satir,column=5).border = thin_border
            sayfa.cell(satir,column=6).border = thin_border
            sayfa.cell(satir,column=7).border = thin_border
            sayfa.cell(satir,column=8).border = thin_border
            sayfa.cell(satir,column=9).border = thin_border
            sayfa.cell(satir,column=10).border = thin_border
            sayfa.cell(satir,column=11).border = thin_border
            sayfa.cell(satir,column=12).border = thin_border
            sayfa.cell(satir,column=13).border = thin_border
            sayfa.cell(satir,column=14).border = thin_border
            sayfa.cell(satir,column=15).border = thin_border



            crates_size.border = thin_border
            crates_size.font = font
            crates_size.alignment  = Alignment(horizontal="center", vertical="center")


            merge_9 = 'E' + str(satir + 1) + ':' +'O' + str(satir + 3)
            sayfa.merge_cells(merge_9)
            sayfa.cell(satir + 1,column=5).border = thin_border
            sayfa.cell(satir + 1,column=6).border = thin_border
            sayfa.cell(satir + 1,column=7).border = thin_border
            sayfa.cell(satir + 1,column=8).border = thin_border
            sayfa.cell(satir + 1,column=9).border = thin_border
            sayfa.cell(satir + 1,column=10).border = thin_border
            sayfa.cell(satir + 1,column=11).border = thin_border
            sayfa.cell(satir + 1,column=12).border = thin_border
            sayfa.cell(satir + 1,column=13).border = thin_border
            sayfa.cell(satir + 1,column=14).border = thin_border
            sayfa.cell(satir + 1,column=15).border = thin_border



            sayfa.cell(satir + 2,column=5).border = thin_border
            sayfa.cell(satir + 2,column=6).border = thin_border
            sayfa.cell(satir + 2,column=7).border = thin_border
            sayfa.cell(satir + 2,column=8).border = thin_border
            sayfa.cell(satir + 2,column=9).border = thin_border
            sayfa.cell(satir + 2,column=10).border = thin_border
            sayfa.cell(satir + 2,column=11).border = thin_border
            sayfa.cell(satir + 2,column=12).border = thin_border
            sayfa.cell(satir + 2,column=13).border = thin_border
            sayfa.cell(satir + 2,column=14).border = thin_border
            sayfa.cell(satir + 2,column=15).border = thin_border

            sayfa.cell(satir + 3,column=5).border = thin_border
            sayfa.cell(satir + 3,column=6).border = thin_border
            sayfa.cell(satir + 3,column=7).border = thin_border
            sayfa.cell(satir + 3,column=8).border = thin_border
            sayfa.cell(satir + 3,column=9).border = thin_border
            sayfa.cell(satir + 3,column=10).border = thin_border
            sayfa.cell(satir + 3,column=11).border = thin_border
            sayfa.cell(satir + 3,column=12).border = thin_border
            sayfa.cell(satir + 3,column=13).border = thin_border
            sayfa.cell(satir + 3,column=14).border = thin_border
            sayfa.cell(satir + 3,column=15).border = thin_border

            satir += 5
            merge_10 = 'A' + str(satir) + ':' +'D' + str(satir)
            sayfa.merge_cells(merge_10)
            container = sayfa.cell(satir,column = 1,value="KONTEYNER İÇİ DAĞILIM")
            sayfa.cell(satir,column=2).border = thin_border
            sayfa.cell(satir,column=3).border = thin_border
            sayfa.cell(satir,column=4).border = thin_border


            container.border = thin_border
            container.font = font
            container.alignment  = Alignment(horizontal="center", vertical="center")


            merge_11 = 'A' + str(satir + 1) + ':' +'D' + str(satir + 3)
            sayfa.merge_cells(merge_11)
            sayfa.cell(satir + 1,column=1).border = thin_border
            sayfa.cell(satir + 1,column=2).border = thin_border
            sayfa.cell(satir + 1,column=3).border = thin_border
            sayfa.cell(satir + 1,column=4).border = thin_border

            sayfa.cell(satir + 2,column=1).border = thin_border
            sayfa.cell(satir + 2,column=2).border = thin_border
            sayfa.cell(satir + 2,column=3).border = thin_border
            sayfa.cell(satir + 2,column=4).border = thin_border

            sayfa.cell(satir + 3,column=1).border = thin_border
            sayfa.cell(satir + 3,column=2).border = thin_border
            sayfa.cell(satir + 3,column=3).border = thin_border
            sayfa.cell(satir + 3,column=4).border = thin_border




            merge_12 = 'E' + str(satir) + ':' +'O' + str(satir)
            sayfa.merge_cells(merge_12)
            notes = sayfa.cell(satir,column = 5,value="YÜKLEME TOPLANTISI NOTLARI")
            sayfa.cell(satir,column=5).border = thin_border
            sayfa.cell(satir,column=6).border = thin_border
            sayfa.cell(satir,column=7).border = thin_border
            sayfa.cell(satir,column=8).border = thin_border
            sayfa.cell(satir,column=9).border = thin_border
            sayfa.cell(satir,column=10).border = thin_border
            sayfa.cell(satir,column=11).border = thin_border
            sayfa.cell(satir,column=12).border = thin_border
            sayfa.cell(satir,column=13).border = thin_border
            sayfa.cell(satir,column=14).border = thin_border
            sayfa.cell(satir,column=15).border = thin_border



            notes.border = thin_border
            notes.font = font
            notes.alignment  = Alignment(horizontal="center", vertical="center")


            merge_13 = 'E' + str(satir + 1) + ':' +'O' + str(satir + 3)
            sayfa.merge_cells(merge_13)
            sayfa.cell(satir + 1,column=5).border = thin_border
            sayfa.cell(satir + 1,column=6).border = thin_border
            sayfa.cell(satir + 1,column=7).border = thin_border
            sayfa.cell(satir + 1,column=8).border = thin_border
            sayfa.cell(satir + 1,column=9).border = thin_border
            sayfa.cell(satir + 1,column=10).border = thin_border
            sayfa.cell(satir + 1,column=11).border = thin_border
            sayfa.cell(satir + 1,column=12).border = thin_border
            sayfa.cell(satir + 1,column=13).border = thin_border
            sayfa.cell(satir + 1,column=14).border = thin_border
            sayfa.cell(satir + 1,column=15).border = thin_border



            sayfa.cell(satir + 2,column=5).border = thin_border
            sayfa.cell(satir + 2,column=6).border = thin_border
            sayfa.cell(satir + 2,column=7).border = thin_border
            sayfa.cell(satir + 2,column=8).border = thin_border
            sayfa.cell(satir + 2,column=9).border = thin_border
            sayfa.cell(satir + 2,column=10).border = thin_border
            sayfa.cell(satir + 2,column=11).border = thin_border
            sayfa.cell(satir + 2,column=12).border = thin_border
            sayfa.cell(satir + 2,column=13).border = thin_border
            sayfa.cell(satir + 2,column=14).border = thin_border
            sayfa.cell(satir + 2,column=15).border = thin_border

            sayfa.cell(satir + 3,column=5).border = thin_border
            sayfa.cell(satir + 3,column=6).border = thin_border
            sayfa.cell(satir + 3,column=7).border = thin_border
            sayfa.cell(satir + 3,column=8).border = thin_border
            sayfa.cell(satir + 3,column=9).border = thin_border
            sayfa.cell(satir + 3,column=10).border = thin_border
            sayfa.cell(satir + 3,column=11).border = thin_border
            sayfa.cell(satir + 3,column=12).border = thin_border
            sayfa.cell(satir + 3,column=13).border = thin_border
            sayfa.cell(satir + 3,column=14).border = thin_border
            sayfa.cell(satir + 3,column=15).border = thin_border

            kitap.save(target_path)
            kitap.close()

            return True

         except Exception as e:
            print('ceki_listesi_excel  Hata : ',str(e))
            return False  
    
    def seleksiyon_listesi_excel(self,data_list):
        try:
        
            source_path = 'excel/sablonlar/seleksiyon_etiket.xlsx'
            target_path = 'excel/dosyalar/seleksiyon_etiket.xlsx'

            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sheet']
            satir = 1
            column = 1
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            for i in data_list:
 

                case_no = sayfa.cell(satir,column=column,value='CASE NO:')
                case_no.border = thin_border
                case_no.font = Font(bold=True)
                case_no.alignment = Alignment(horizontal="center", vertical="center")
                case_merge = 'B' + str(satir) + ':' + 'D' + str(satir)
                sayfa.merge_cells(case_merge)
                case = sayfa.cell(satir,column=column + 1,value=i['KasaNo'])
                sayfa.cell(satir,column=column + 2).border = thin_border
                sayfa.cell(satir,column=column + 3).border = thin_border
                case.border = thin_border
                case.alignment = Alignment(horizontal="center", vertical="center")
                case.font = Font(size=36,bold=True)

                status = sayfa.cell(satir + 1,column=column,value='STATUS: ')
                status.border = thin_border
                status.alignment = Alignment(horizontal="center", vertical="center")
                status.font = Font(bold=True)
                status_merge = 'B' + str(satir + 1) + ':' + 'D' + str(satir + 1)
                sayfa.merge_cells(status_merge)

                stock = sayfa.cell(satir + 1,column=column + 1,value='STOK')
                sayfa.cell(satir+1,column=column + 2).border = thin_border
                sayfa.cell(satir+1,column=column + 3).border = thin_border
                stock.border = thin_border
                stock.alignment = Alignment(horizontal="center", vertical="center")
                stock.font = Font(size=36,bold=True)

                destination = sayfa.cell(satir,column = column + 4,value = 'DESTINATION')
                destination.border = thin_border
                destination.alignment = Alignment(horizontal="center", vertical="center")
                destination.font = Font(bold=True)
                country = sayfa.cell(satir + 1,column = column + 4,value = 'TR')
                country.border = thin_border
                country.alignment = Alignment(horizontal="center", vertical="center")
                country.font = Font(size=14,bold=True)

                product_name = i['KategoriAdi'] + ' / ' + i['UrunAdi'] + ' / ' + i['YuzeyIslemAdi'] + ' / ' + str(i['En']) + 'x' + str(i['Boy']) + 'x' +  str(i['Kenar']) + ' / ' + str(i['Adet']) + ' pcs / ' + str(i['Miktar']) + ' ' + i['UrunBirimAdi']
                product_merge = 'A' + str(satir + 2) + ':' + 'E' + str(satir + 7)
                sayfa.merge_cells(product_merge)
                product = sayfa.cell(satir + 2,column = column,value=product_name)
                product.border = thin_border
                product.alignment = Alignment(horizontal="center", vertical="center",wrapText=True)
                product.font = Font(size=24,bold=True)
                sayfa.cell(satir + 2,column = column).border = thin_border
                sayfa.cell(satir + 3,column = column).border = thin_border
                sayfa.cell(satir + 4,column = column).border = thin_border
                sayfa.cell(satir + 5,column = column).border = thin_border
                sayfa.cell(satir + 6,column = column).border = thin_border
                sayfa.cell(satir + 7,column = column).border = thin_border

                sayfa.cell(satir + 2,column = column+4).border = thin_border
                sayfa.cell(satir + 3,column = column+4).border = thin_border
                sayfa.cell(satir + 4,column = column+4).border = thin_border
                sayfa.cell(satir + 5,column = column+4).border = thin_border
                sayfa.cell(satir + 6,column = column+4).border = thin_border
                sayfa.cell(satir + 7,column = column+4).border = thin_border

                sayfa.cell(satir + 7,column = column+1).border = thin_border
                sayfa.cell(satir + 7,column = column+2).border = thin_border
                sayfa.cell(satir + 7,column = column+3).border = thin_border

                satir += 9


            kitap.save(target_path)
            kitap.close()
            return True

                

                

        except Exception as e:
            print('seleksiyon etiket çıktı hata',str(e))
            return False
    
    def finance_excel_custom(self,data_list):
        try:
            source_path = 'excel/sablonlar/finance_detail_custom.xlsx'
            target_path = 'excel/dosyalar/finance_detail_custom.xlsx'

            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            sayfa2= kitap['Sayfa2']
            satir = 1
            satir2 = 1
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            po = sayfa.cell(satir,column=1,value='Po')
            po.border = thin_border
            po.font = Font(bold=True)
            order_date = sayfa.cell(satir,column=2,value='Order Date')
            order_date.border = thin_border
            order_date.font = Font(bold=True)

            shipped_date = sayfa.cell(satir,column=3,value='Shipped Date')
            shipped_date.border = thin_border
            shipped_date.font = Font(bold=True)

            status = sayfa.cell(satir,column=4,value='Status')
            status.border = thin_border
            status.font = Font(bold=True)

            order_total = sayfa.cell(satir,column=5,value='Order Total')
            order_total.border = thin_border
            order_total.font = Font(bold=True)

            payment_received = sayfa.cell(satir,column=6,value='Payment Received')
            payment_received.border = thin_border
            payment_received.font = Font(bold=True)


            balance = sayfa.cell(satir,column=7,value='Balanced')
            balance.border = thin_border
            balance.font = Font(bold=True)

            pre_payment = sayfa.cell(satir,column=8,value='Prepayment')
            pre_payment.border = thin_border
            pre_payment.font = Font(bold=True)
            satir += 1
            for po in data_list['po']:
                sayfa.cell(satir,column=1,value=po['SiparisNo']).border = thin_border
                sayfa.cell(satir,column=2,value=self._dateConvert(po['SiparisTarihi'])).border = thin_border
                sayfa.cell(satir,column=3,value=self._dateConvert(po['YuklemeTarihi'])).border = thin_border
                sayfa.cell(satir,column=4,value=po['Durum']).border = thin_border
                sayfa.cell(satir,column=5,value=self._formatControl(po['OrderTotal'])).border = thin_border
                sayfa.cell(satir,column=6,value=self._formatControl(po['Paid'])).border = thin_border
                sayfa.cell(satir,column=7,value=self._formatControl(po['Balanced'])).border = thin_border
                sayfa.cell(satir,column=8,value=self._formatControl(po['Pesinat'])).border = thin_border
                satir += 1
            
            paid_date = sayfa2.cell(satir2,column=1,value="Date")
            paid_date.border = thin_border
            paid_date.font = Font(bold = True)
            paid = sayfa2.cell(satir2,column=2,value="Paid")
            paid.border = thin_border
            paid.font = Font(bold = True)
            satir2 += 1
            for paid in data_list['paid']:
                sayfa2.cell(satir2,column=1,value=self._dateConvert(paid['Tarih'])).border = thin_border
                sayfa2.cell(satir2,column=2,value=self._formatControl(paid['Paid'])).border = thin_border
                satir2 += 1


            kitap.save(target_path)
            kitap.close()
            return True



        except Exception as e:
            print('finance_excel_custom',str(e))
            return False
    
    def _dateConvert(self,date):
        if(date == None or date == '' or date == ' '):
            return ''
        else:
            if(date):
                _year,_month,_date = str(date).split('-')
                newDate = datetime.datetime(int(_year),int(_month),int(_date[0:2]))
                year = newDate.strftime('%Y')
                month = newDate.strftime('%m')
                day = newDate.strftime('%d')
                return str(day) + '-' + str(month) + '-' + str(year)
            else:
                return ''

    def _formatControl(self,val):
        if(val > -8 and val < 8):
            return 0
        else:
            return val
    def __noneControl(self,value):
        if value == None:
            return 0
        else:
            return float(value)
    def reports_strips_excel(self,strips):
        try:
            source_path = 'excel/sablonlar/reports_mekmer_strips.xlsx'
            target_path = 'excel/dosyalar/reports_mekmer_strips.xlsx'

            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 1
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            po = sayfa.cell(satir,column=1,value='Tarih')
            po.border = thin_border
            po.font = Font(bold=True)
            order_date = sayfa.cell(satir,column=2,value='Tedarikçi Adı')
            order_date.border = thin_border
            order_date.font = Font(bold=True)

            shipped_date = sayfa.cell(satir,column=3,value='Ocak Adı')
            shipped_date.border = thin_border
            shipped_date.font = Font(bold=True)

            status = sayfa.cell(satir,column=4,value='Strip Adı')
            status.border = thin_border
            status.font = Font(bold=True)

            order_total = sayfa.cell(satir,column=5,value='Strip (M2)')
            order_total.border = thin_border
            order_total.font = Font(bold=True)

            payment_received = sayfa.cell(satir,column=6,value='Strip ($)')
            payment_received.border = thin_border
            payment_received.font = Font(bold=True)


            balance = sayfa.cell(satir,column=7,value='Strip Toplam ($)')
            balance.border = thin_border
            balance.font = Font(bold=True)

            pre_payment = sayfa.cell(satir,column=8,value='En')
            pre_payment.border = thin_border
            pre_payment.font = Font(bold=True)

            pre_payment = sayfa.cell(satir,column=9,value='Boy')
            pre_payment.border = thin_border
            pre_payment.font = Font(bold=True)

            pre_payment = sayfa.cell(satir,column=10,value='Adet')
            pre_payment.border = thin_border
            pre_payment.font = Font(bold=True)
            satir += 1
            for strip in strips:
                sayfa.cell(satir,column=1,value=self._dateConvert(strip['Date'])).border = thin_border
                sayfa.cell(satir,column=2,value=strip['SupplierName']).border = thin_border
                sayfa.cell(satir,column=3,value=strip['QuarryName']).border = thin_border
                sayfa.cell(satir,column=4,value=strip['StripName']).border = thin_border
                sayfa.cell(satir,column=5,value=strip['StripM2']).border = thin_border
                sayfa.cell(satir,column=6,value=strip['StripPrice']).border = thin_border
                sayfa.cell(satir,column=7,value=strip['StripCost']).border = thin_border
                sayfa.cell(satir,column=8,value=strip['StripWidth']).border = thin_border
                sayfa.cell(satir,column=9,value=strip['StripHeight']).border = thin_border
                sayfa.cell(satir,column=10,value=strip['StripPiece']).border = thin_border



            
            


            kitap.save(target_path)
            kitap.close()
            return True



        except Exception as e:
            print('finance_excel_custom',str(e))
            return False
    def reports_moloz_excel(self,moloz):
        try:
            source_path = 'excel/sablonlar/reports_mekmer_moloz.xlsx'
            target_path = 'excel/dosyalar/reports_mekmer_moloz.xlsx'

            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 1
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            po = sayfa.cell(satir,column=1,value='Tarih')
            po.border = thin_border
            po.font = Font(bold=True)
            order_date = sayfa.cell(satir,column=2,value='Tedarikçi Adı')
            order_date.border = thin_border
            order_date.font = Font(bold=True)

            shipped_date = sayfa.cell(satir,column=3,value='Ocak Adı')
            shipped_date.border = thin_border
            shipped_date.font = Font(bold=True)

            status = sayfa.cell(satir,column=4,value='Strip Adı')
            status.border = thin_border
            status.font = Font(bold=True)

            order_total = sayfa.cell(satir,column=5,value='Tonaj')
            order_total.border = thin_border
            order_total.font = Font(bold=True)

            payment_received = sayfa.cell(satir,column=6,value='Fiyat (Tl)')
            payment_received.border = thin_border
            payment_received.font = Font(bold=True)


            balance = sayfa.cell(satir,column=7,value='Fiyat ($)')
            balance.border = thin_border
            balance.font = Font(bold=True)

            currency = sayfa.cell(satir,column=8,value='Kur ($)')
            currency.border = thin_border
            currency.font = Font(bold=True)

            pre_payment = sayfa.cell(satir,column=9,value='Toplam (Tl)')
            pre_payment.border = thin_border
            pre_payment.font = Font(bold=True)




            satir += 1
            for strip in moloz:
                sayfa.cell(satir,column=1,value=self._dateConvert(strip['Date'])).border = thin_border
                sayfa.cell(satir,column=2,value=strip['SupplierName']).border = thin_border
                sayfa.cell(satir,column=3,value=strip['QuarryName']).border = thin_border
                sayfa.cell(satir,column=4,value=strip['StripName']).border = thin_border
                sayfa.cell(satir,column=5,value=strip['Ton']).border = thin_border
                sayfa.cell(satir,column=6,value=strip['PriceTl']).border = thin_border
                sayfa.cell(satir,column=7,value=strip['PriceUsd']).border = thin_border
                sayfa.cell(satir,column=8,value=strip['Currency']).border = thin_border
                sayfa.cell(satir,column=9,value=strip['Total']).border = thin_border




            kitap.save(target_path)
            kitap.close()
            return True



        except Exception as e:
            print('finance_excel_custom',str(e))
            return False

    def customer_mekmer_excel(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            source_path = 'excel/sablonlar/mekmar_customer.xlsx'
            target_path = 'excel/dosyalar/mekmar_customer.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 2
            for item in data:
                sayfa.cell(satir,column=1,value=item['ID']).border = thin_border
                sayfa.cell(satir,column=2,value=item['FirmaAdi']).border = thin_border
                sayfa.cell(satir,column=3,value=item['Unvan']).border = thin_border
                sayfa.cell(satir,column=4,value=item['Adres']).border = thin_border
                sayfa.cell(satir,column=5,value=item['UlkeAdi']).border = thin_border
                sayfa.cell(satir,column=6,value=item['Marketing']).border = thin_border
                sayfa.cell(satir,column=7,value=item['MailAdresi']).border = thin_border
                sayfa.cell(satir,column=8,value=item['SatisciAdi']).border = thin_border
                sayfa.cell(satir,column=9,value=item['Temsilci']).border = thin_border   
                satir += 1

            kitap.save(target_path)
            kitap.close()
            return True



        except Exception as e:
            print('customer_mekmer_excel hata',e)
            return False

    def selection_excel_output(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            source_path = 'excel/sablonlar/selection.xlsx'
            target_path = 'excel/dosyalar/selection.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 2
            
            for item in data:
                sayfa.cell(satir,column=1,value=item['KasaNo']).border = thin_border
                sayfa.cell(satir,column=2,value=item['OcakAdi']).border = thin_border
                sayfa.cell(satir,column=3,value=item['FirmaAdi']).border = thin_border
                sayfa.cell(satir,column=4,value=item['KategoriAdi']).border = thin_border
                sayfa.cell(satir,column=5,value=item['UrunAdi']).border = thin_border
                sayfa.cell(satir,column=6,value=item['YuzeyIslemAdi']).border = thin_border
                sayfa.cell(satir,column=7,value=item['En']).border = thin_border
                sayfa.cell(satir,column=8,value=item['Boy']).border = thin_border
                sayfa.cell(satir,column=9,value=item['Kenar']).border = thin_border 
                sayfa.cell(satir,column=10,value=item['KutuAdet']).border = thin_border   
                sayfa.cell(satir,column=11,value=item['KutuIciAdet']).border = thin_border   
                sayfa.cell(satir,column=12,value=item['Adet']).border = thin_border   
                sayfa.cell(satir,column=13,value=item['Miktar']).border = thin_border   
                sayfa.cell(satir,column=14,value=item['UrunBirimAdi']).border = thin_border   
                sayfa.cell(satir,column=15,value=item['SiparisAciklama']).border = thin_border   
                sayfa.cell(satir,column=16,value=item['Aciklama']).border = thin_border   
                sayfa.cell(satir,column=17,value=item['Fason']).border = thin_border   


                satir += 1

            kitap.save(target_path)
            kitap.close()
            return True



        except Exception as e:
            print('selection product hata',e)
            return False


    def selection_excel_mail_output(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            source_path = 'excel/sablonlar/selection.xlsx'
            target_path = 'excel/dosyalar/selection.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 2
            
            for item in data:
                
                sayfa.cell(satir,column=1,value=item.KasaNo).border = thin_border
                sayfa.cell(satir,column=2,value=item.OcakAdi).border = thin_border
                sayfa.cell(satir,column=3,value=item.FirmaAdi).border = thin_border
                sayfa.cell(satir,column=4,value=item.KategoriAdi).border = thin_border
                sayfa.cell(satir,column=5,value=item.UrunAdi).border = thin_border
                sayfa.cell(satir,column=6,value=item.YuzeyIslemAdi).border = thin_border
                sayfa.cell(satir,column=7,value=item.En).border = thin_border
                sayfa.cell(satir,column=8,value=item.Boy).border = thin_border
                sayfa.cell(satir,column=9,value=item.Kenar).border = thin_border 
                sayfa.cell(satir,column=10,value=item.KutuAdet).border = thin_border   
                sayfa.cell(satir,column=11,value=item.KutuIciAdet).border = thin_border   
                sayfa.cell(satir,column=12,value=item.Adet).border = thin_border   
                sayfa.cell(satir,column=13,value=item.Miktar).border = thin_border   
                sayfa.cell(satir,column=14,value=item.UrunBirimAdi).border = thin_border   
                sayfa.cell(satir,column=15,value=item.SiparisAciklama).border = thin_border   
                sayfa.cell(satir,column=16,value=item.Aciklama).border = thin_border   
                sayfa.cell(satir,column=17,value=item.Fason).border = thin_border   


                satir += 1

            kitap.save(target_path)
            kitap.close()
            return True



        except Exception as e:
            print('selection product hata',e)
            return False


    def production_excel_list(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/production_excel.xlsx'
            target_path = 'excel/dosyalar/production_excel.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 2
            amount_total = 0
            piece_total = 0



            for item in data:
                sayfa.cell(satir,column=1,value=self._dateConvert(item['Tarih'])).border = thin_border
                sayfa.cell(satir,column=2,value=item['FirmaAdi']).border = thin_border
                sayfa.cell(satir,column=3,value=item['KategoriAdi']).border = thin_border
                sayfa.cell(satir,column=4,value=item['KasaNo']).border = thin_border
                sayfa.cell(satir,column=5,value=item['UrunAdi']).border = thin_border
                sayfa.cell(satir,column=6,value=item['OcakAdi']).border = thin_border
                sayfa.cell(satir,column=7,value=item['YuzeyIslemAdi']).border = thin_border
                sayfa.cell(satir,column=8,value=item['En']).border = thin_border
                sayfa.cell(satir,column=9,value=item['Boy']).border = thin_border
                sayfa.cell(satir,column=10,value=item['Kenar']).border = thin_border
                sayfa.cell(satir,column=11,value=item['Miktar']).border = thin_border
                sayfa.cell(satir,column=12,value=item['Adet']).border = thin_border
                sayfa.cell(satir,column=13,value=item['BirimAdi']).border = thin_border
                sayfa.cell(satir,column=14,value=item['SiparisAciklama']).border = thin_border
                sayfa.cell(satir,column=15,value=item['Aciklama']).border = thin_border


                amount_total += self.__noneControl(item['Miktar'])
                piece_total += self.__noneControl(item['Adet'])




                satir += 1
            total_cell = sayfa.cell(satir,column=1,value='Total')
            total_cell.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell.border = thin_border
            total_cell.font = font_bold
            total_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            sayfa.cell(satir,column=2,value='').border = thin_border
            sayfa.cell(satir,column=3,value='').border = thin_border
            sayfa.cell(satir,column=4,value='').border = thin_border
            sayfa.cell(satir,column=5,value='').border = thin_border
            sayfa.cell(satir,column=6,value='').border = thin_border
            sayfa.cell(satir,column=7,value='').border = thin_border
            sayfa.cell(satir,column=8,value='').border = thin_border
            sayfa.cell(satir,column=9,value='').border = thin_border
            sayfa.cell(satir,column=10,value='').border = thin_border


            amount_cell = sayfa.cell(satir,column=11,value=amount_total)
            piece_cell = sayfa.cell(satir,column=12,value=piece_total)
            amount_cell.font = font_bold
            amount_cell.border = thin_border
            amount_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            piece_cell.border = thin_border
            piece_cell.font = font_bold
            piece_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")



            other_on_uc_cell = sayfa.cell(satir,column=13,value='') 
            other_on_uc_cell.border = thin_border
            other_on_uc_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            other_on_dort_cell = sayfa.cell(satir,column=14,value='')
            other_on_dort_cell.border = thin_border
            other_on_dort_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            other_on_bes_cell = sayfa.cell(satir,column=15,value='')
            other_on_bes_cell.border = thin_border
            other_on_bes_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            merge_2 = 'A' + str(satir) + ':' + 'J' + str(satir)
            sayfa.merge_cells(merge_2)





            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False

    def mine_excel_list(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/mine_excel.xlsx'
            target_path = 'excel/dosyalar/mine_excel.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 2
            m2_total = 0
            piece_total = 0
            mt_total = 0
            crate_total = 0
            for item in data:
                sayfa.cell(satir,column=1,value=item['OcakAdi']).border = thin_border
                sayfa.cell(satir,column=2,value=item['M2']).border = thin_border
                sayfa.cell(satir,column=3,value=item['MT']).border = thin_border
                sayfa.cell(satir,column=4,value=item['Adet']).border = thin_border
                sayfa.cell(satir,column=5,value=item['KasaAdedi']).border = thin_border
                m2_total += self.__noneControl(item['M2'])
                piece_total += self.__noneControl(item['Adet'])
                mt_total += self.__noneControl(item['MT'])
                crate_total += self.__noneControl(item['KasaAdedi'])






                satir += 1
            total_cell = sayfa.cell(satir,column=1,value='Total')
            total_cell.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell.border = thin_border
            total_cell.font = font_bold
            total_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            amount_cell = sayfa.cell(satir,column=2,value=m2_total)
            piece_cell = sayfa.cell(satir,column=3,value=piece_total)
            m2_cell = sayfa.cell(satir,column=4,value=mt_total)
            crate_cell = sayfa.cell(satir,column=5,value=crate_total)

            amount_cell.font = font_bold
            amount_cell.border = thin_border
            amount_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            piece_cell.border = thin_border
            piece_cell.font = font_bold
            piece_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            m2_cell.border = thin_border
            m2_cell.font = font_bold
            m2_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            crate_cell.border = thin_border
            crate_cell.font = font_bold
            crate_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False

    def loading_excel_list(self,data):
        try:
            print(data)
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/loading_excel_list.xlsx'
            target_path = 'excel/dosyalar/loading_excel_list.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 2
            fob_total = 0
            ddp_total = 0

            for item in data:
               
                sayfa.cell(satir,column=1,value=self._dateConvert(item['YuklemeTarihi'])).border = thin_border
                sayfa.cell(satir,column=2,value=item['MusteriAdi']).border = thin_border
                sayfa.cell(satir,column=3,value=item['SiparisNo']).border = thin_border
                sayfa.cell(satir,column=4,value=item['Fob']).border = thin_border
                sayfa.cell(satir,column=5,value=item['Dtp']).border = thin_border
                fob_total += self.__noneControl(item['Fob'])
                ddp_total += self.__noneControl(item['Dtp'])
                satir += 1

            total_cell = sayfa.cell(satir,column=1,value='Total')
            total_cell.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell.border = thin_border
            total_cell.font = font_bold
            total_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            merge_2 = 'A' + str(satir) + ':' + 'C' + str(satir)
            sayfa.merge_cells(merge_2)

            col_1 = sayfa.cell(satir,column=4,value=fob_total)
            col_2 = sayfa.cell(satir,column=5,value=ddp_total)

            col_1.border = thin_border
            col_1.font = font_bold
            col_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            col_2.border = thin_border
            col_2.font = font_bold
            col_2.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False

    def forwarding_excel_list(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/forwarding_excel.xlsx'
            target_path = 'excel/dosyalar/forwarding_excel.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 2
            amount = 0
            peaceincrate = 0
            boxincrate = 0
            total_price =0


            for item in data:
               
                sayfa.cell(satir,column=1,value=self._dateConvert(item['Tarih'])).border = thin_border
                sayfa.cell(satir,column=2,value=item['FirmaAdi']).border = thin_border
                sayfa.cell(satir,column=3,value=item['TedarikciAdi']).border = thin_border
                sayfa.cell(satir,column=4,value=item['UrunKartId']).border = thin_border
                sayfa.cell(satir,column=5,value=item['KasaNo']).border = thin_border
                sayfa.cell(satir,column=6,value=item['OcakAdi']).border = thin_border
                sayfa.cell(satir,column=7,value=item['KategoriAdi']).border = thin_border
                sayfa.cell(satir,column=8,value=item['UrunAdi']).border = thin_border
                sayfa.cell(satir,column=9,value=item['YuzeyIslemAdi']).border = thin_border
                sayfa.cell(satir,column=10,value=item['En']).border = thin_border
                sayfa.cell(satir,column=11,value=item['Boy']).border = thin_border
                sayfa.cell(satir,column=12,value=item['Kenar']).border = thin_border
                sayfa.cell(satir,column=13,value=item['KutuAdet']).border = thin_border
                sayfa.cell(satir,column=14,value=item['Adet']).border = thin_border
                sayfa.cell(satir,column=15,value=item['Miktar']).border = thin_border
                sayfa.cell(satir,column=16,value=item['BirimAdi']).border = thin_border
                sayfa.cell(satir,column=17,value=item['SiparisAciklama']).border = thin_border
                sayfa.cell(satir,column=18,value=item['Aciklama']).border = thin_border
                sayfa.cell(satir,column=19,value=item['BirimFiyat']).border = thin_border
                sayfa.cell(satir,column=20,value=item['Toplam']).border = thin_border

                amount += self.__noneControl(item['Miktar'])
                peaceincrate += self.__noneControl(item['Adet'])
                boxincrate += self.__noneControl(item['KutuAdet'])
                total_price += self.__noneControl(item['Toplam'])
                satir += 1

            total_cell = sayfa.cell(satir,column=1,value='Total')
            total_cell.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell.border = thin_border
            total_cell.font = font_bold
            total_cell.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            merge_2 = 'A' + str(satir) + ':' + 'L' + str(satir)
            sayfa.merge_cells(merge_2)

            col_1 = sayfa.cell(satir,column=13,value=boxincrate)
            col_2 = sayfa.cell(satir,column=14,value=peaceincrate)
            col_3 = sayfa.cell(satir,column=15,value=amount)
            col_4 = sayfa.cell(satir,column=20,value=total_price)


            col_1.border = thin_border
            col_1.font = font_bold
            col_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            col_2.border = thin_border
            col_2.font = font_bold
            col_2.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            col_3.border = thin_border
            col_3.font = font_bold
            col_3.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            col_4.border = thin_border
            col_4.font = font_bold
            col_4.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False


    def orders_by_country(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/orders_by_country.xlsx'
            target_path = 'excel/dosyalar/orders_by_country.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            satir = 3
            header_2024 = sayfa.cell(1,column=1,value='2024')
            header_2023=sayfa.cell(1,column=5,value='2023')
            header_2022 = sayfa.cell(1,column=9,value='2022')
            header_2021=sayfa.cell(1,column=13,value='2021')
            header_2020=sayfa.cell(1,column=17,value='2020')
            header_2019 =sayfa.cell(1,column=21,value='2019')


            sayfa.cell(2,column=1,value='Country').border = thin_border
            sayfa.cell(2,column=2,value='FOB').border = thin_border
            sayfa.cell(2,column=3,value='DDP').border = thin_border

            sayfa.cell(2,column=5,value='Country').border = thin_border
            sayfa.cell(2,column=6,value='FOB').border = thin_border
            sayfa.cell(2,column=7,value='DDP').border = thin_border

            sayfa.cell(2,column=9,value='Country').border = thin_border
            sayfa.cell(2,column=10,value='FOB').border = thin_border
            sayfa.cell(2,column=11,value='DDP').border = thin_border


            sayfa.cell(2,column=13,value='Country').border = thin_border
            sayfa.cell(2,column=14,value='FOB').border = thin_border
            sayfa.cell(2,column=15,value='DDP').border = thin_border

            sayfa.cell(2,column=17,value='Country').border = thin_border
            sayfa.cell(2,column=18,value='FOB').border = thin_border
            sayfa.cell(2,column=19,value='DDP').border = thin_border

            
            sayfa.cell(2,column=21,value='Country').border = thin_border
            sayfa.cell(2,column=22,value='FOB').border = thin_border
            sayfa.cell(2,column=23,value='DDP').border = thin_border




            header_2024.border = thin_border
            header_2023.border = thin_border
            header_2022.border = thin_border
            header_2021.border = thin_border
            header_2020.border = thin_border
            header_2019.border = thin_border

            header_2024.font = font_bold
            header_2023.font = font_bold
            header_2022.font = font_bold
            header_2021.font = font_bold
            header_2020.font = font_bold
            header_2019.font = font_bold

            header_2024.alignment  = Alignment(horizontal="center", vertical="center")
            header_2023.alignment  = Alignment(horizontal="center", vertical="center")
            header_2022.alignment  = Alignment(horizontal="center", vertical="center")
            header_2021.alignment  = Alignment(horizontal="center", vertical="center")
            header_2020.alignment  = Alignment(horizontal="center", vertical="center")
            header_2019.alignment  = Alignment(horizontal="center", vertical="center")

            header_2024.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2023.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2022.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2021.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2020.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2019.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            merge_2024 = 'A' + str(1) + ':' + 'C' + str(1)
            merge_2023 = 'E' + str(1) + ':' + 'G' + str(1)
            merge_2022 = 'I' + str(1) + ':' + 'K' + str(1)
            merge_2021 = 'M' + str(1) + ':' + 'O' + str(1)
            merge_2020 = 'Q' + str(1) + ':' + 'S' + str(1)
            merge_2019 = 'U' + str(1) + ':' + 'W' + str(1)


            


            fob_total_1 = 0
            ddp_total_1 = 0

            for item in data['this_year']:
               
                sayfa.cell(satir,column=1,value=item['UlkeAdi']).border = thin_border
                sayfa.cell(satir,column=2,value=item['Fob']).border = thin_border
                sayfa.cell(satir,column=3,value=item['ddp']).border = thin_border

                fob_total_1 += self.__noneControl(item['Fob'])
                ddp_total_1 += self.__noneControl(item['ddp'])
                satir += 1

            total_cell_1 = sayfa.cell(satir,column=1,value='Total')
            total_cell_1.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell_1.border = thin_border
            total_cell_1.font = font_bold
            total_cell_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            col_1 = sayfa.cell(satir,column=2,value=fob_total_1)
            col_2 = sayfa.cell(satir,column=3,value=ddp_total_1)

            col_1.border = thin_border
            col_1.font = font_bold
            col_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            col_2.border = thin_border
            col_2.font = font_bold
            col_2.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")




            fob_total_2 = 0
            ddp_total_2 = 0
            satir_2 = 3
            for item in data['one_year_ago']:
               
                sayfa.cell(satir_2,column=5,value=item['UlkeAdi']).border = thin_border
                sayfa.cell(satir_2,column=6,value=item['Fob']).border = thin_border
                sayfa.cell(satir_2,column=7,value=item['ddp']).border = thin_border

                fob_total_2 += self.__noneControl(item['Fob'])
                ddp_total_2 += self.__noneControl(item['ddp'])
                satir_2 += 1

            total_cell_2 = sayfa.cell(satir_2,column=5,value='Total')
            total_cell_2.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell_2.border = thin_border
            total_cell_2.font = font_bold
            total_cell_2.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            col_3 = sayfa.cell(satir_2,column=6,value=fob_total_2)
            col_4 = sayfa.cell(satir_2,column=7,value=ddp_total_2)
            
            col_3.border = thin_border
            col_3.font = font_bold
            col_3.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            col_4.border = thin_border
            col_4.font = font_bold
            col_4.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            



            fob_total_3 = 0
            ddp_total_3 = 0
            satir_3 = 3
            for item in data['two_year_ago']:
               
                sayfa.cell(satir_3,column=9,value=item['UlkeAdi']).border = thin_border
                sayfa.cell(satir_3,column=10,value=item['Fob']).border = thin_border
                sayfa.cell(satir_3,column=11,value=item['ddp']).border = thin_border

                fob_total_3 += self.__noneControl(item['Fob'])
                ddp_total_3 += self.__noneControl(item['ddp'])
                satir_3 += 1

            total_cell_3 = sayfa.cell(satir_3,column=9,value='Total')
            total_cell_3.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell_3.border = thin_border
            total_cell_3.font = font_bold
            total_cell_3.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            col_5 = sayfa.cell(satir_3,column=10,value=fob_total_3)
            col_6 = sayfa.cell(satir_3,column=11,value=ddp_total_3)
            
            col_5.border = thin_border
            col_5.font = font_bold
            col_5.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            col_6.border = thin_border
            col_6.font = font_bold
            col_6.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")




            fob_total_4 = 0
            ddp_total_4 = 0
            satir_4 = 3
            for item in data['three_year_ago']:
               
                sayfa.cell(satir_4,column=13,value=item['UlkeAdi']).border = thin_border
                sayfa.cell(satir_4,column=14,value=item['Fob']).border = thin_border
                sayfa.cell(satir_4,column=15,value=item['ddp']).border = thin_border

                fob_total_4 += self.__noneControl(item['Fob'])
                ddp_total_4 += self.__noneControl(item['ddp'])
                satir_4 += 1

            total_cell_4 = sayfa.cell(satir_4,column=13,value='Total')
            total_cell_4.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell_4.border = thin_border
            total_cell_4.font = font_bold
            total_cell_4.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            col_7 = sayfa.cell(satir_4,column=14,value=fob_total_4)
            col_8 = sayfa.cell(satir_4,column=15,value=ddp_total_4)
            
            col_7.border = thin_border
            col_7.font = font_bold
            col_7.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            col_8.border = thin_border
            col_8.font = font_bold
            col_8.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")




            fob_total_5 = 0
            ddp_total_5 = 0
            satir_5 = 3
            for item in data['four_year_ago']:
               
                sayfa.cell(satir_5,column=17,value=item['UlkeAdi']).border = thin_border
                sayfa.cell(satir_5,column=18,value=item['Fob']).border = thin_border
                sayfa.cell(satir_5,column=19,value=item['ddp']).border = thin_border

                fob_total_5 += self.__noneControl(item['Fob'])
                ddp_total_5 += self.__noneControl(item['ddp'])
                satir_5 += 1

            total_cell_5 = sayfa.cell(satir_5,column=17,value='Total')
            total_cell_5.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell_5.border = thin_border
            total_cell_5.font = font_bold
            total_cell_5.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            col_7 = sayfa.cell(satir_5,column=18,value=fob_total_5)
            col_8 = sayfa.cell(satir_5,column=19,value=ddp_total_5)
            
            col_7.border = thin_border
            col_7.font = font_bold
            col_7.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            col_8.border = thin_border
            col_8.font = font_bold
            col_8.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            fob_total_6 = 0
            ddp_total_6 = 0
            satir_6 = 3
            for item in data['five_year_ago']:
               
                sayfa.cell(satir_6,column=21,value=item['UlkeAdi']).border = thin_border
                sayfa.cell(satir_6,column=22,value=item['Fob']).border = thin_border
                sayfa.cell(satir_6,column=23,value=item['ddp']).border = thin_border

                fob_total_6 += self.__noneControl(item['Fob'])
                ddp_total_6 += self.__noneControl(item['ddp'])
                satir_6 += 1

            total_cell_6 = sayfa.cell(satir_6,column=21,value='Total')
            total_cell_6.alignment  = Alignment(horizontal="center", vertical="center")
            total_cell_6.border = thin_border
            total_cell_6.font = font_bold
            total_cell_6.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")


            col_9 = sayfa.cell(satir_6,column=22,value=fob_total_6)
            col_10 = sayfa.cell(satir_6,column=23,value=ddp_total_6)
            
            col_9.border = thin_border
            col_9.font = font_bold
            col_9.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            col_10.border = thin_border
            col_10.font = font_bold
            col_10.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")





            sayfa.merge_cells(merge_2024)
            sayfa.merge_cells(merge_2023)
            sayfa.merge_cells(merge_2022)
            sayfa.merge_cells(merge_2021)
            sayfa.merge_cells(merge_2020)
            sayfa.merge_cells(merge_2019)

            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False

    def gu_forwarding(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/gu_excel.xlsx'
            target_path = 'excel/dosyalar/gu_excel.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            
            header_2024 = sayfa.cell(1,column=1,value='2024')
            header_2023=sayfa.cell(1,column=5,value='2023')
            header_2022 = sayfa.cell(1,column=9,value='2022')
            header_2021=sayfa.cell(1,column=13,value='2021')
            header_2020=sayfa.cell(1,column=17,value='2020')
            header_2019 =sayfa.cell(1,column=21,value='2019')
            header_2018 =sayfa.cell(1,column=25,value='2018')
            header_2017 =sayfa.cell(1,column=29,value='2017')
            header_2016 =sayfa.cell(1,column=33,value='2016')
            header_2015 =sayfa.cell(1,column=37,value='2015')
            header_2014 =sayfa.cell(1,column=41,value='2014')




            cell_2024_1 = sayfa.cell(2,column=1,value='Month')
            cell_2024_2 =sayfa.cell(2,column=2,value='FOB')
            cell_2024_3 =sayfa.cell(2,column=3,value='DDP')

            cell_2023_1 =sayfa.cell(2,column=5,value='Month')
            cell_2023_2 =sayfa.cell(2,column=6,value='FOB')
            cell_2023_3 =sayfa.cell(2,column=7,value='DDP')

            cell_2022_1 =sayfa.cell(2,column=9,value='Month')
            cell_2022_2 =sayfa.cell(2,column=10,value='FOB')
            cell_2022_3 =sayfa.cell(2,column=11,value='DDP')


            cell_2021_1 =sayfa.cell(2,column=13,value='Month')
            cell_2021_2 =sayfa.cell(2,column=14,value='FOB')
            cell_2021_3 =sayfa.cell(2,column=15,value='DDP')

            cell_2020_1 =sayfa.cell(2,column=17,value='Month')
            cell_2020_2 =sayfa.cell(2,column=18,value='FOB')
            cell_2020_3 =sayfa.cell(2,column=19,value='DDP')

            
            cell_2019_1 =sayfa.cell(2,column=21,value='Month')
            cell_2019_2 =sayfa.cell(2,column=22,value='FOB')
            cell_2019_3 =sayfa.cell(2,column=23,value='DDP')

            cell_2018_1 =sayfa.cell(2,column=25,value='Month')
            cell_2018_2 =sayfa.cell(2,column=26,value='FOB')
            cell_2018_3 =sayfa.cell(2,column=27,value='DDP')

            cell_2017_1 =sayfa.cell(2,column=29,value='Month')
            cell_2017_2 =sayfa.cell(2,column=30,value='FOB')
            cell_2017_3 =sayfa.cell(2,column=31,value='DDP')

            cell_2016_1 =sayfa.cell(2,column=33,value='Month')
            cell_2016_2 =sayfa.cell(2,column=34,value='FOB')
            cell_2016_3 =sayfa.cell(2,column=35,value='DDP')

            cell_2015_1 =sayfa.cell(2,column=37,value='Month')
            cell_2015_2 =sayfa.cell(2,column=38,value='FOB')
            cell_2015_3 =sayfa.cell(2,column=39,value='DDP')

            
            cell_2014_1 =sayfa.cell(2,column=41,value='Month')
            cell_2014_2 =sayfa.cell(2,column=42,value='FOB')
            cell_2014_3 =sayfa.cell(2,column=43,value='DDP')

            cell_2024_1.border = thin_border
            cell_2024_1.font = font_bold
            cell_2024_1.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2024_2.border = thin_border
            cell_2024_2.font = font_bold
            cell_2024_2.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2024_3.border = thin_border
            cell_2024_3.font = font_bold
            cell_2024_3.alignment  = Alignment(horizontal="center", vertical="center")


            cell_2023_1.border = thin_border
            cell_2023_1.font = font_bold
            cell_2023_1.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2023_2.border = thin_border
            cell_2023_2.font = font_bold
            cell_2023_2.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2023_3.border = thin_border
            cell_2023_3.font = font_bold
            cell_2023_3.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2022_1.border = thin_border
            cell_2022_1.font = font_bold
            cell_2022_1.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2022_2.border = thin_border
            cell_2022_2.font = font_bold
            cell_2022_2.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2022_3.border = thin_border
            cell_2022_3.font = font_bold
            cell_2022_3.alignment  = Alignment(horizontal="center", vertical="center")


            cell_2021_1.border = thin_border
            cell_2021_1.font = font_bold
            cell_2021_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2021_2.border = thin_border
            cell_2021_2.font = font_bold
            cell_2021_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2021_3.border = thin_border
            cell_2021_3.font = font_bold
            cell_2021_3.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2020_1.border = thin_border
            cell_2020_1.font = font_bold
            cell_2020_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2020_2.border = thin_border
            cell_2020_2.font = font_bold
            cell_2020_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2020_3.border = thin_border
            cell_2020_3.font = font_bold
            cell_2020_3.alignment  = Alignment(horizontal="center", vertical="center")


            cell_2019_1.border = thin_border
            cell_2019_1.font = font_bold
            cell_2019_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2019_2.border = thin_border
            cell_2019_2.font = font_bold
            cell_2019_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2019_3.border = thin_border
            cell_2019_3.font = font_bold
            cell_2019_3.alignment  = Alignment(horizontal="center", vertical="center")

            
            cell_2018_1.border = thin_border
            cell_2018_1.font = font_bold
            cell_2018_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2018_2.border = thin_border
            cell_2018_2.font = font_bold
            cell_2018_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2018_3.border = thin_border
            cell_2018_3.font = font_bold
            cell_2018_3.alignment  = Alignment(horizontal="center", vertical="center")

                        
            cell_2017_1.border = thin_border
            cell_2017_1.font = font_bold
            cell_2017_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2017_2.border = thin_border
            cell_2017_2.font = font_bold
            cell_2017_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2017_3.border = thin_border
            cell_2017_3.font = font_bold
            cell_2017_3.alignment  = Alignment(horizontal="center", vertical="center")

            cell_2016_1.border = thin_border
            cell_2016_1.font = font_bold
            cell_2016_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2016_2.border = thin_border
            cell_2016_2.font = font_bold
            cell_2016_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2016_3.border = thin_border
            cell_2016_3.font = font_bold
            cell_2016_3.alignment  = Alignment(horizontal="center", vertical="center")
            
            
            cell_2015_1.border = thin_border
            cell_2015_1.font = font_bold
            cell_2015_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2015_2.border = thin_border
            cell_2015_2.font = font_bold
            cell_2015_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2015_3.border = thin_border
            cell_2015_3.font = font_bold
            cell_2015_3.alignment  = Alignment(horizontal="center", vertical="center")

                        
            cell_2014_1.border = thin_border
            cell_2014_1.font = font_bold
            cell_2014_1.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2014_2.border = thin_border
            cell_2014_2.font = font_bold
            cell_2014_2.alignment  = Alignment(horizontal="center", vertical="center")
            cell_2014_3.border = thin_border
            cell_2014_3.font = font_bold
            cell_2014_3.alignment  = Alignment(horizontal="center", vertical="center")








            header_2024.border = thin_border
            header_2023.border = thin_border
            header_2022.border = thin_border
            header_2021.border = thin_border
            header_2020.border = thin_border
            header_2019.border = thin_border
            header_2018.border = thin_border
            header_2017.border = thin_border
            header_2016.border = thin_border
            header_2015.border = thin_border
            header_2014.border = thin_border


            header_2024.font = font_bold
            header_2023.font = font_bold
            header_2022.font = font_bold
            header_2021.font = font_bold
            header_2020.font = font_bold
            header_2019.font = font_bold
            header_2018.font = font_bold
            header_2017.font = font_bold
            header_2016.font = font_bold
            header_2015.font = font_bold
            header_2014.font = font_bold

            header_2024.alignment  = Alignment(horizontal="center", vertical="center")
            header_2023.alignment  = Alignment(horizontal="center", vertical="center")
            header_2022.alignment  = Alignment(horizontal="center", vertical="center")
            header_2021.alignment  = Alignment(horizontal="center", vertical="center")
            header_2020.alignment  = Alignment(horizontal="center", vertical="center")
            header_2019.alignment  = Alignment(horizontal="center", vertical="center")
            header_2018.alignment  = Alignment(horizontal="center", vertical="center")
            header_2017.alignment  = Alignment(horizontal="center", vertical="center")
            header_2016.alignment  = Alignment(horizontal="center", vertical="center")
            header_2015.alignment  = Alignment(horizontal="center", vertical="center")
            header_2014.alignment  = Alignment(horizontal="center", vertical="center")

            header_2024.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2023.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2022.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2021.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2020.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2019.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2018.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2017.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2016.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2015.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
            header_2014.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

            merge_2024 = 'A' + str(1) + ':' + 'C' + str(1)
            merge_2023 = 'E' + str(1) + ':' + 'G' + str(1)
            merge_2022 = 'I' + str(1) + ':' + 'K' + str(1)
            merge_2021 = 'M' + str(1) + ':' + 'O' + str(1)
            merge_2020 = 'Q' + str(1) + ':' + 'S' + str(1)
            merge_2019 = 'U' + str(1) + ':' + 'W' + str(1)
            merge_2018 = 'Y' + str(1) + ':' + 'AA' + str(1)
            merge_2017 = 'AC' + str(1) + ':' + 'AE' + str(1)
            merge_2016 = 'AG' + str(1) + ':' + 'AI' + str(1)
            merge_2015 = 'AK' + str(1) + ':' + 'AM' + str(1)
            merge_2014 = 'AO' + str(1) + ':' + 'AQ' + str(1)
            
            sayfa.merge_cells(merge_2024)
            sayfa.merge_cells(merge_2023)
            sayfa.merge_cells(merge_2022)
            sayfa.merge_cells(merge_2021)
            sayfa.merge_cells(merge_2020)
            sayfa.merge_cells(merge_2019)
            sayfa.merge_cells(merge_2018)
            sayfa.merge_cells(merge_2017)
            sayfa.merge_cells(merge_2016)
            sayfa.merge_cells(merge_2015)
            sayfa.merge_cells(merge_2014)




            



            column = 1
            
            for item in data:
                satir = 3
                fob_total_1 = 0
                ddp_total_1 = 0
                for i in item['data']:
                    sayfa.cell(satir,column=column,value=i['Ay']).border = thin_border
                    sayfa.cell(satir,column=column+1,value=i['Fob']).border = thin_border
                    sayfa.cell(satir,column=column+2,value=i['Ddp']).border = thin_border

                    fob_total_1 += self.__noneControl(i['Fob'])
                    ddp_total_1 += self.__noneControl(i['Ddp'])
                    satir += 1

                total_cell_1 = sayfa.cell(satir,column=column,value='Total')
                total_cell_1.alignment  = Alignment(horizontal="center", vertical="center")
                total_cell_1.border = thin_border
                total_cell_1.font = font_bold
                total_cell_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")



                col_1 = sayfa.cell(satir,column=column+1,value=fob_total_1)
                col_2 = sayfa.cell(satir,column=column+2,value=ddp_total_1)

                col_1.border = thin_border
                col_1.font = font_bold
                col_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
                col_2.border = thin_border
                col_2.font = font_bold
                col_2.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

                column += 4



   


        






            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False


    def gu_supplier_cost(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/supplier_cost.xlsx'
            target_path = 'excel/dosyalar/supplier_cost.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            column = 1
            for item in data:
                satir = 1
                header = sayfa.cell(satir,column=column,value=item['year'])
                satir += 1
                cell_2024_1 = sayfa.cell(satir,column=column,value='Supplier')
                cell_2024_2 = sayfa.cell(satir,column=column+1,value='Total')
                satir += 1
                cell_2024_1.border = thin_border
                cell_2024_1.font = font_bold
                cell_2024_1.alignment  = Alignment(horizontal="center", vertical="center")
                cell_2024_2.border = thin_border
                cell_2024_2.font = font_bold
                cell_2024_2.alignment  = Alignment(horizontal="center", vertical="center")

                header.border = thin_border
                header.font = font_bold
                header.alignment  = Alignment(horizontal="center", vertical="center")
                header.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
                # merge = 'A' + str(1) + ':' + 'C' + str(1)
                # sayfa.merge_cells(merge)
                


                total = 0
                for i in item['data']:
                    sayfa.cell(satir,column=column,value=i['FirmaAdi']).border = thin_border
                    sayfa.cell(satir,column=column+1,value=i['Total']).border = thin_border

                    total += self.__noneControl(i['Total'])
                    satir += 1

                total_cell_1 = sayfa.cell(satir,column=column,value='Total')
                total_cell_1.alignment  = Alignment(horizontal="center", vertical="center")
                total_cell_1.border = thin_border
                total_cell_1.font = font_bold
                total_cell_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")



                col_1 = sayfa.cell(satir,column=column+1,value=total)

                col_1.border = thin_border
                col_1.font = font_bold
                col_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

                column += 3

            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False
    
    def gu_continents(self,data):
        try:
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
            
            font_bold = Font(name='Calibri',
                size=13,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
            source_path = 'excel/sablonlar/continents.xlsx'
            target_path = 'excel/dosyalar/continents.xlsx'
            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap['Sayfa1']
            column = 1
            for item in data:
                satir = 1
                header = sayfa.cell(satir,column=column,value=item['year'])
                satir += 1
                cell_2024_1 = sayfa.cell(satir,column=column,value='Continents')
                cell_2024_2 = sayfa.cell(satir,column=column+1,value='Fob')
                cell_2024_3 = sayfa.cell(satir,column=column+2,value='Ddp')

                satir += 1
                cell_2024_1.border = thin_border
                cell_2024_1.font = font_bold
                cell_2024_1.alignment  = Alignment(horizontal="center", vertical="center")
                cell_2024_2.border = thin_border
                cell_2024_2.font = font_bold
                cell_2024_2.alignment  = Alignment(horizontal="center", vertical="center")
                cell_2024_3.border = thin_border
                cell_2024_3.font = font_bold
                cell_2024_3.alignment  = Alignment(horizontal="center", vertical="center")

                header.border = thin_border
                header.font = font_bold
                header.alignment  = Alignment(horizontal="center", vertical="center")
                header.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
                # merge = 'A' + str(1) + ':' + 'C' + str(1)
                # sayfa.merge_cells(merge)
                


                fob = 0
                ddp = 0

                for i in item['data']:
                    sayfa.cell(satir,column=column,value=i['Continent']).border = thin_border
                    sayfa.cell(satir,column=column+1,value=i['Fob']).border = thin_border
                    sayfa.cell(satir,column=column+2,value=i['Ddp']).border = thin_border


                    fob += self.__noneControl(i['Fob'])
                    ddp += self.__noneControl(i['Ddp'])

                    satir += 1

                total_cell_1 = sayfa.cell(satir,column=column,value='Total')
                total_cell_1.alignment  = Alignment(horizontal="center", vertical="center")
                total_cell_1.border = thin_border
                total_cell_1.font = font_bold
                total_cell_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")



                col_1 = sayfa.cell(satir,column=column+1,value=fob)
                col_2 = sayfa.cell(satir,column=column+2,value=ddp)


                col_1.border = thin_border
                col_1.font = font_bold
                col_1.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")
                col_2.border = thin_border
                col_2.font = font_bold
                col_2.fill = PatternFill(start_color="defb00", end_color="defb00", fill_type = "solid")

                column += 4

            kitap.save(target_path)
            kitap.close()
            return True
        except Exception as e:
            print('production_excel hata',e)
            return False

    def mailGonder(self,data):
        mail = smtplib.SMTP("mail.mekmar.com",587)
        mail.ehlo()
        mail.starttls()
        mail.login("goz@mekmar.com", "_bwt64h-3SR_-G2O")
        mesaj = MIMEMultipart()
        mesaj["From"] = "goz@mekmar.com"           # Gönderen
        mesaj["Subject"] = f"{str(datetime.datetime.now().year)}   -   {str(datetime.datetime.now().month)}  ' ayı Seleksiyon Listesi'"
        mesaj["To"] = 'bilgiislem@mekmar.com'
        body = """
                <table>
                    <tr>
                        <th>Kasa No</th>
                        <th>Ocak</th>
                        <th>Firma Adı</th>
                        <th>Kategori</th>
                        <th>Ürün</th>
                        <th>Yüzey</th>
                        <th>En</th>
                        <th>Boy</th>
                        <th>Kenar</th>
                        <th>Kutu</th>
                        <th>Kutu içi Adet</th>
                        <th>Adet</th>
                        <th>Miktar</th>
                        <th>Birim</th>
                        <th>Po</th>
                        <th>Açıklama</th>
                        <th>Fason</th>
                    </tr>
                    
        """

        for item in data:
            body+=f"""
            <tr style="border-bottom:1px solid gray;">
                <td style="border-bottom:1px solid gray;">{item.KasaNo}</td>
                <td style="border-bottom:1px solid gray;">{item.OcakAdi}</td>
                <td style="border-bottom:1px solid gray;">{item.FirmaAdi}</td>
                <td style="border-bottom:1px solid gray;">{item.KategoriAdi}</td>
                <td style="border-bottom:1px solid gray;">{item.UrunAdi}</td>
                <td style="border-bottom:1px solid gray;">{item.YuzeyIslemAdi}</td>
                <td style="border-bottom:1px solid gray;">{item.En}</td>
                <td style="border-bottom:1px solid gray;">{item.Boy}</td>
                <td style="border-bottom:1px solid gray;">{item.Kenar}</td>
                <td style="border-bottom:1px solid gray;">{item.KutuAdet}</td>
                <td style="border-bottom:1px solid gray;">{item.KutuIciAdet}</td>
                <td style="border-bottom:1px solid gray;">{item.Adet}</td>
                <td style="border-bottom:1px solid gray;">{item.Miktar}</td>
                <td style="border-bottom:1px solid gray;">{item.UrunBirimAdi}</td>
                <td style="border-bottom:1px solid gray;">{item.SiparisAciklama}</td>
                <td style="border-bottom:1px solid gray;">{item.Aciklama}</td>
                <td style="border-bottom:1px solid gray;">{item.Fason}</td>
            </tr>



            
            """
        
        body += "</table>"
            
            

        part2 = MIMEText(body, 'html')
        mesaj.attach(part2)

 
 

        
        mail.sendmail(mesaj["From"], mesaj["To"], mesaj.as_string())
        print("Mail başarılı bir şekilde gönderildi.")
        mail.close()
        return True




        

            


class SiparisCekiListesiApi(Resource):

    def post(self):

        data_list = request.get_json()

        islem = ExcellCiktiIslem()

        result = islem.ceki_listesi_excel(data_list)

        return jsonify({'status' : result})

    def get(self):

        excel_path = 'excel/dosyalar/ceki_listesi.xlsx'

        return send_file(excel_path,as_attachment=True)
    
class SeleksiyonUrunEtiketApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.seleksiyon_listesi_excel(data)
        return jsonify({'status':status})
    
    def get(self):

        excel_path = 'excel/dosyalar/seleksiyon_etiket.xlsx'

        return send_file(excel_path,as_attachment=True)

class FinanceTestExcelCustomApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.finance_excel_custom(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/finance_detail_custom.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsStripsExcelApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.reports_strips_excel(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/reports_mekmer_strips.xlsx'
        return send_file(excel_path,as_attachment=True)


class ReportsMolozExcelApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.reports_moloz_excel(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/reports_mekmer_moloz.xlsx'
        return send_file(excel_path,as_attachment=True)

class CustomerMekmerExcelApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.customer_mekmer_excel(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/mekmar_customer.xlsx'
        return send_file(excel_path,as_attachment=True)

class SelectionExcelApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.selection_excel_output(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/selection.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsProductionApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.production_excel_list(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/production_excel.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsProductMineApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.mine_excel_list(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/mine_excel.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsProductLoadingApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.loading_excel_list(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/loading_excel_list.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsProductForwardingApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.forwarding_excel_list(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/forwarding_excel.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsOrdersByCountryApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.orders_by_country(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/orders_by_country.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsGuForwardingApi(Resource):
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.gu_forwarding(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/gu_excel.xlsx'
        return send_file(excel_path,as_attachment=True)
    
class ReportsGuSupplierCostApi(Resource):
    
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.gu_supplier_cost(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/supplier_cost.xlsx'
        return send_file(excel_path,as_attachment=True)

class ReportsGuContinentsApi(Resource):
    
    def post(self):
        data = request.get_json()
        excel = ExcellCiktiIslem()
        status = excel.gu_continents(data)
        return jsonify({'status':status})
    def get(self):
        excel_path = 'excel/dosyalar/continents.xlsx'
        return send_file(excel_path,as_attachment=True)


class ReportsProductionMailSendApi(Resource):
    def get(self):
        productionSql = """
                select top 1 YEAR(Tarih) as Year,MONTH(Tarih) as Month,DAY(Tarih) as Day from UretimTB
                where TedarikciID in (1,123)
                order by ID desc
        """
    

        sql = """
                select u.OzelMiktar,
    u.Fason,
    u.Duzenleyen,
	u.Kasalayan,
	u.UrunKartId,
	urb.ID as UrunBirimId,
	urb.BirimAdi as UrunBirimAdi,
	u.UretimTurID,
	u.ID,
	u.Tarih,
	u.KasaNo,
	k.KategoriAdi,
	k.ID as KategoriID,
	uo.OcakAdi,
	uo.ID as OcakId,
	ur.UrunAdi,
	ur.ID as UrunId,
	yk.YuzeyIslemAdi,
	yk.ID as YuzeyId,
	ol.ID as OlcuId,
	ol.En,
	ol.Boy,
	ol.Kenar,
	u.KutuAdet,
	u.KutuIciAdet,
	u.Miktar,
	u.Kutu,
	u.Bagli,
	u.SiparisAciklama,
	u.Aciklama,
	u.TedarikciID,
	t.FirmaAdi,
	u.Bulunamadi,
	u.Disarda,
	u.Adet from UretimTB u 
	inner join TedarikciTB t on t.ID = u.TedarikciID 
	inner join UrunBirimTB ub on ub.ID = u.UrunBirimID 
	inner join UrunOcakTB uo on uo.ID = u.UrunOcakID 
	inner join UretimTurTB ut on ut.ID = u.UretimTurID 
	inner join UrunKartTB uk on uk.ID = u.UrunKartID 
	inner join KategoriTB k on k.ID = uk.KategoriID 
	inner join UrunlerTB ur on ur.ID = uk.UrunID 
	inner join YuzeyKenarTB yk on yk.ID = uk.YuzeyID 
	inner join OlculerTB ol on ol.ID = uk.OlcuID 
	inner join UrunBirimTB urb on urb.ID = u.UrunBirimID 
	where u.UrunDurumID=1 and u.TedarikciID in (1,123) and u.Bulunamadi != 1
							
	order by KasaNo desc
        """
        sqlIslem = SqlConnect().data
        excel = ExcellCiktiIslem()
        production = sqlIslem.getList(productionSql)
        productionYear = production[0].Year
        productionMonth = production[0].Month
        productionDay = production[0].Day

        

        date = datetime.datetime.now()
        year = date.year
        month = date.month
        lastMonthDay = calendar.monthrange(year,month)
        nowDay = date.day
        lastDay = lastMonthDay[1]

        day = datetime.datetime(int(year),int(month),int(lastDay)).strftime('%A')
        if(day == 'Saturday'):
            lastDay -= 1
        if(day == 'Sunday'):
            lastDay -= 2
        if(lastDay == nowDay):
            results = sqlIslem.getList(sql)
            status = sqlIslem.getList("select Status from SeleksiyonSendMailTB")
            if(status[0].Status == False):
                # islem = excel.selection_excel_mail_output(results)
                if(len(results)) > 0:
                    mailStatus = excel.mailGonder(results)
                    if(mailStatus):
                        sqlIslem.update_insert("update SeleksiyonSendMailTB SET Status=?",(1))
            else:
                if(lastDay >= 1 and lastDay <= 5):
                    sqlIslem.update_insert("update SeleksiyonSendMailTB SET Status=?",(0))





        


api.add_resource(SiparisCekiListesiApi, '/excel/check/list', methods=['GET','POST'])
api.add_resource(MaliyetRaporIslemApi,'/maliyet/listeler/maliyetListesi/<int:yil>/<int:ay>',methods=['GET'])
api.add_resource(MaliyetRaporIslemYilApi,'/maliyet/listeler/maliyetListesi/<int:yil>',methods=['GET'])
api.add_resource(FinanceTestListApi,'/finance/reports/test',methods=['GET'])
api.add_resource(FinanceTestListExcelApi,'/finance/reports/test/excel',methods=['GET','POST'])
api.add_resource(FinanceTestListExcelApiFilter,'/finance/reports/test/excel/mekmer',methods=['GET','POST'])

api.add_resource(UretimExcelCiktiApi,'/siparisler/dosyalar/uretimExcelCikti',methods=['POST','GET'])
api.add_resource(MkRaporlariExcelApi,'/raporlar/listeler/mkraporlari/excel',methods=['GET','POST'])
api.add_resource(MkRaporlariApi,'/raporlar/listeler/mkraporlari/<int:year>',methods=['GET'])
api.add_resource(StokRaporExcelApi,'/raporlar/listeler/stokRaporExcelListe',methods=['GET','POST'])
api.add_resource(MaliyetRaporExcelApi, '/maliyet/dosyalar/maliyetRaporExcelListe', methods=['GET','POST'])
api.add_resource(CurrencyApi,'/finance/doviz/liste/<string:yil>/<string:ay>/<string:gun>',methods=['GET'])
api.add_resource(CurrencyUsdToEuroApi,'/finance/doviz/liste/usd/to/euro/<string:yil>/<string:ay>/<string:gun>',methods=['GET'])
api.add_resource(CurrencyEuroToTlApi,'/finance/doviz/liste/euro/to/tl/<string:yil>/<string:ay>/<string:gun>',methods=['GET'])




api.add_resource(FinanceTestListFilterApi,'/finance/reports/test/filter',methods=['GET'])
api.add_resource(FinanceTestDetailFilterApi,'/finance/po/list/mekmer/<int:customer>',methods=['GET'])

api.add_resource(FinanceTestDetailFilterMonthApi,'/finance/po/list/mekmer/month/<int:month>',methods=['GET'])



api.add_resource(FinanceTestPaidFilterApi,'/finance/po/paid/mekmer/save',methods=['POST'])
api.add_resource(FinanceTestPoPaidListFilterApi,'/finance/po/paid/list/mekmer/<string:po>',methods=['GET'])

api.add_resource(FinanceTestListFilterMekmerAllApi,'/finance/reports/mekmer/all',methods=['GET'])

api.add_resource(FinanceTestListFilterPoApi,'/finance/mekmar/po/paid/save',methods=['POST'])

api.add_resource(FinanceTestExcelCustomApi,'/finance/mekmar/excel/custom',methods=['GET','POST'])


api.add_resource(GuReportsSellerAndOperationOrdersExcelApi,'/gu/reports/seller/operation/orders',methods=['GET','POST'])
api.add_resource(GuReportsSellerAndOperationForwardingExcelApi,'/gu/reports/seller/operation/forwarding',methods=['GET','POST'])
api.add_resource(SeleksiyonUrunEtiketApi,'/seleksiyon/etiket/excel',methods=['GET','POST'])
api.add_resource(ReportsStripsExcelApi,'/reports/mekmer/strips/excel',methods=['GET','POST'])
api.add_resource(ReportsMolozExcelApi,'/reports/mekmer/moloz/excel',methods=['GET','POST'])


api.add_resource(CreditCardCostYearApi,'/reports/mekmar/ayo/credit/card/<int:year>',methods=['GET'])




api.add_resource(AyoCostExcelApi,'/reports/mekmar/ayo/cost/excel',methods=['GET','POST'])


api.add_resource(CustomerMekmerExcelApi,'/customer/mekmar/excel',methods=['GET','POST'])


api.add_resource(SelectionExcelApi,'/siparisler/dosyalar/seleksiyon/excel/output',methods=['GET','POST'])


#Excels

api.add_resource(ReportsProductionApi,'/reports/excel/production',methods=['GET','POST'])
api.add_resource(ReportsProductMineApi,'/reports/excel/mine',methods=['GET','POST'])
api.add_resource(ReportsProductLoadingApi,'/reports/excel/loading',methods=['GET','POST'])
api.add_resource(ReportsProductForwardingApi,'/reports/excel/forwarding',methods=['GET','POST'])
api.add_resource(ReportsOrdersByCountryApi,'/maliyet/dosyalar/countries',methods=['GET','POST'])
api.add_resource(ReportsGuForwardingApi,'/reports/gu/forwarding',methods=['GET','POST'])
api.add_resource(ReportsGuSupplierCostApi,'/reports/excel/supplier/cost',methods=['GET','POST'])
api.add_resource(ReportsGuContinentsApi,'/maliyet/dosyalar/continent',methods=['GET','POST'])



api.add_resource(ReportsProductionMailSendApi,'/reports/production/send/mail',methods=['GET'])




if __name__ == '__main__':
    app.run(port=5000,debug=True)